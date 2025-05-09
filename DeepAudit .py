import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import requests
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter
import json
from pathlib import Path
import configparser
from datetime import datetime
from tkinter import scrolledtext
import traceback
import threading
from queue import Queue, Empty
import re
import os
import queue
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from tkinter import ttk, filedialog, messagebox, scrolledtext, simpledialog
import tkinter as tk


class CodeAuditApp:
    def __init__(self, root):
        self.event_queue = Queue()
        self.root = root
        # API验证错误提示状态标志
        self.api_validation_error_shown = False
        self.validation_lock = threading.Lock()  # 添加线程锁
        self.root.title("DeepAudit 代码审计工具")
        self.root.geometry("1280x800")

        # 初始化漏洞列表
        self.vulnerabilities = {}

        # 初始化分析状态
        self.auto_analysis_cancelled = False
        self.auto_analysis_paused = False
        self.api_validation_error_shown = False  # 移除validation_lock的使用

        # 初始化进度相关变量
        self.progress_var = tk.IntVar(value=0)
        self.progress_bar = None
        self.files_to_analyze = []

        # 添加API验证状态标志，用于记录API是否已验证成功
        self.api_validated = False

        # 启动事件处理循环
        self.root.after(100, self.process_event_queue)

        # 新增配置初始化（关键修复）
        self.config = configparser.ConfigParser()
        self.config_path = Path.cwd() / 'config.ini'
        if not self.config_path.exists():
            self.config['DEFAULT'] = {
                'API_KEY': '',
                'API_ENDPOINT': 'https://api.deepseek.com/v1/chat/completions',
                'TIMEOUT': '30'
            }
            self._save_config()
        else:
            self.config.read(self.config_path, encoding='utf-8')

        # 修正API终端地址（添加/v1/chat/completions路径）
        self.api_endpoint = self.config['DEFAULT'].get('API_ENDPOINT', 'https://api.deepseek.com/v1/chat/completions')
        self.api_key = self.config.get('DEFAULT', 'API_KEY', fallback='')
        print(f"[DEBUG] 最终API终端: {self.api_endpoint}")

        # 初始化项目路径为当前目录
        self.project_path = Path.cwd()  # 新增默认路径初始化

        # 初始化核心组件
        self._init_ui()
        self._bind_events()

        # 添加全局ID计数器
        self.vuln_id_counter = 1

        # 添加控制自动分析的状态变量
        self.auto_analysis_paused = False
        self.auto_analysis_cancelled = False
        self.auto_analysis_thread = None

        # 修改初始化顺序，立即开始配置初始化
        self.delayed_config_init()

        # 不再自动验证API密钥，只在需要时验证
        self.status_bar.config(text="就绪")

    def _init_configuration(self):
        """统一配置初始化"""
        self.config = configparser.ConfigParser()
        self.config_path = Path("config.ini")

        # 初始化API相关配置
        self.api_key = self.config.get('DEFAULT', 'API_KEY', fallback='')
        self.api_endpoint = self.config.get('DEFAULT', 'API_ENDPOINT',
                                            fallback='https://api.deepseek.com/v1/chat/completions')

        # 初始化日志路径
        self.log_file = self.project_path / 'error.log'
        self.api_log_file = self.project_path / 'api.log'

        # 支持的语言类型
        self.supported_langs = {
            '.php': 'php',
            '.java': 'java',
            '.js': 'javascript',
            '.html': 'html',
            '.xml': 'xml'
        }

        # 风险等级颜色
        self.severity_colors = {
            '高危': '#ff4444',
            '中危': '#ffa500',
            '低危': '#4CAF50'
        }

        # 延迟加载非必要配置
        self.root.after(500, self.load_full_config)

    def _init_ui(self):
        """界面组件初始化"""
        # 主框架
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 工具栏
        self._init_toolbar(main_frame)

        # 创建可调整大小的面板容器
        self.main_paned = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        self.main_paned.pack(fill=tk.BOTH, expand=True)

        # 初始化左侧文件树和右侧面板
        self._init_file_tree(self.main_paned)
        self._init_right_panel(self.main_paned)

        # 状态栏
        self._init_status_bar()
        # 延迟设置分隔条位置，确保窗口已完全渲染
        self.root.after(100, self._set_initial_pane_position)

    def _set_initial_pane_position(self):
        """设置PanedWindow的初始分隔条位置"""
        try:
            # 确保窗口已经渲染
            self.root.update_idletasks()

            # 获取窗口宽度
            window_width = self.root.winfo_width()

            # 如果窗口宽度有效，设置分隔条位置为窗口宽度的1/4
            if window_width > 100:  # 确保窗口宽度有效
                left_panel_width = int(window_width / 5)  # 左侧面板占1/4
                self.main_paned.sashpos(0, left_panel_width)
            else:
                # 如果窗口宽度无效，使用固定宽度
                self.main_paned.sashpos(0, 200)  # 默认左侧面板宽度为250像素

            # 记录初始宽度，以便窗口调整时保持比例
            self.initial_left_width = self.main_paned.sashpos(0)

        except Exception as e:
            self.log_error(f"设置分隔条位置失败: {str(e)}")
            # 出错时使用默认值
            self.main_paned.sashpos(0, 200)

    def export_vulnerabilities(self):
        """将漏洞结果导出为Excel文档"""
        if not self.result_tree.get_children():
            messagebox.showwarning("导出失败", "没有可导出的漏洞信息")
            return

        try:
            # 生成带时间戳的文件名
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            default_filename = f"漏洞报告_{timestamp}.xlsx"

            # 获取保存文件路径
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx")],
                initialfile=default_filename
            )

            if not filepath:
                return  # 用户取消了保存

            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            if ws is None:  # 检查ws是否为None
                ws = wb.create_sheet()

            # 设置工作表标题
            if hasattr(ws, 'title'):
                ws.title = "漏洞扫描报告"

            # 设置列宽 - 使用更合适的宽度单位，openpyxl中的宽度单位约为Excel中的0.14倍
            column_widths = {
                'A': 4,  # ID
                'B': 12,  # 漏洞类型
                'C': 12,  # 风险等级
                'D': 50,  # 文件路径
                'E': 60,  # 漏洞描述
                'F': 40,  # 风险点
                'G': 40,  # Payload
                'H': 60,  # 修复建议
                'I': 10,  # 行号
            }

            # 强制设置列宽并确保应用
            if hasattr(ws, 'column_dimensions'):
                for col, width in column_widths.items():
                    # 直接设置列宽，不检查列是否存在
                    ws.column_dimensions[col].width = width
                    # 设置自动换行以确保内容完整显示
                    ws.column_dimensions[col].bestFit = True

            # 设置表头样式
            header_font = Font(name='微软雅黑', bold=True, size=11)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 设置边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 添加表头
            headers = ['ID', '漏洞类型', '风险等级', '文件路径', '漏洞描述', '风险点', 'Payload', '修复建议', '行号']
            for col_num, header in enumerate(headers, 1):
                if hasattr(ws, 'cell'):
                    cell = ws.cell(row=1, column=col_num)
                    if cell is not None:
                        cell.value = header
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border

            # 添加基本信息
            if hasattr(ws, 'merge_cells'):
                ws.merge_cells('A1:I1')

            if hasattr(ws, 'cell'):
                title_cell = ws.cell(row=1, column=1)
                if title_cell is not None:
                    title_cell.value = "漏洞扫描报告"
                    title_cell.font = Font(name='微软雅黑', bold=True, size=14)
                    title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # 添加报告信息
            info_row = 2
            if hasattr(ws, 'merge_cells') and hasattr(ws, 'cell'):
                ws.merge_cells(f'A{info_row}:I{info_row}')
                cell = ws.cell(row=info_row, column=1)
                if cell is not None:
                    cell.value = f"生成时间: {time.strftime('%Y-%m-%d %H:%M:%S')}"

                info_row += 1
                ws.merge_cells(f'A{info_row}:I{info_row}')
                cell = ws.cell(row=info_row, column=1)
                if cell is not None:
                    cell.value = f"项目路径: {self.project_path}"

                info_row += 1
                ws.merge_cells(f'A{info_row}:I{info_row}')
                cell = ws.cell(row=info_row, column=1)
                if cell is not None:
                    # 计算实际漏洞总数
                    actual_vuln_count = len(self.result_tree.get_children())
                    cell.value = f"发现漏洞总数: {actual_vuln_count}"

            # 添加表头（在信息行之后）
            header_row = info_row + 2
            for col_num, header in enumerate(headers, 1):
                if hasattr(ws, 'cell'):
                    cell = ws.cell(row=header_row, column=col_num)
                    if cell is not None:
                        cell.value = header
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border

            # 填充数据
            row_num = header_row + 1
            for item in self.result_tree.get_children():
                values = self.result_tree.item(item, 'values')

                # 确保values有足够的元素
                values = list(values) + [''] * (9 - len(values))

                # 添加数据行
                if hasattr(ws, 'cell'):
                    # 设置行高为自动调整
                    if hasattr(ws, 'row_dimensions'):
                        ws.row_dimensions[row_num].height = None  # 设置为None让Excel自动调整行高

                    for col_num, value in enumerate(values, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        if cell is not None:
                            # 确保行号列正确显示
                            if col_num == 1:
                                # ID列保持原样
                                cell.value = str(value) if value is not None else ""
                            elif col_num == 9:
                                # 行号列（第9列）
                                line_numbers = values[8] if len(values) > 8 and values[8] else "N/A"
                                cell.value = str(line_numbers)
                            else:
                                cell.value = str(value) if value is not None else ""
                            # 优化单元格对齐方式和自动换行设置
                            cell.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
                            cell.border = thin_border

                            # 根据风险等级设置颜色
                            if col_num == 3:  # 风险等级列
                                if value == "高危":
                                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                                elif value == "中危":
                                    cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                                elif value == "低危":
                                    cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

                row_num += 1

            # 保存文件
            wb.save(filepath)
            messagebox.showinfo("导出成功", f"漏洞报告已保存到：\n{filepath}")

        except Exception as e:
            self.log_error(f"导出失败: {str(e)}")
            messagebox.showerror("导出失败", f"导出过程中发生错误：\n{str(e)}")

    def auto_analyze(self):
        """自动分析项目中所有文件"""
        # 检查API密钥是否已配置
        if not self.api_key:
            messagebox.showwarning("警告", "API密钥未配置，请先在设置中配置API密钥")
            return False

        # 如果API尚未验证过，则进行验证
        if not hasattr(self, 'api_validated') or not self.api_validated:
            # 显示验证中的状态
            self.status_bar.config(text="正在验证API密钥...")
            self.root.config(cursor="wait")  # 更改鼠标指针为等待状态

            # 直接验证API密钥，避免异步验证导致的卡顿
            try:
                # 由于_validate_api_key已简化，直接调用并获取结果
                result = self._validate_api_key()

                # 恢复鼠标指针
                self.root.config(cursor="")

                if result:
                    self.api_validated = True
                    self.status_bar.config(text="API密钥验证成功，开始分析...")
                    # 继续执行自动分析流程
                else:
                    messagebox.showerror("API验证失败", "API密钥验证失败，请检查API密钥是否正确。")
                    self.status_bar.config(text="API验证失败")
                    self.btn_auto_analyze.config(text="自动分析")
                    self.btn_analyze.config(state=tk.NORMAL)
                    self.api_validated = False
                    return False
            except Exception as e:
                # 处理验证过程中的异常
                self.root.config(cursor="")
                self.status_bar.config(text="API验证出错")
                messagebox.showerror("API验证错误", f"验证过程中发生错误：{str(e)}")
                self.btn_auto_analyze.config(text="自动分析")
                self.btn_analyze.config(state=tk.NORMAL)
                return False

        # 初始化状态
        self._reset_analysis_state()
        self.auto_analysis_cancelled = False
        self.auto_analysis_paused = False
        self.api_validation_error_shown = False  # 移除validation_lock的使用

        # 更新按钮状态并立即刷新UI
        self.btn_auto_analyze.config(text="取消分析")
        self.btn_analyze.config(state=tk.DISABLED)
        # self.btn_pause_resume.config(state=tk.NORMAL)

        # 强制更新UI，确保按钮文本立即显示
        self.root.update_idletasks()

        # 获取项目中所有文件
        all_files = []
        for root, dirs, files in os.walk(self.project_path):
            for file in files:
                file_path = Path(root) / file
                if file_path.suffix in self.supported_langs:
                    all_files.append(file_path)

        # 初始化进度条
        self.progress['maximum'] = len(all_files)
        self.progress['value'] = 0
        self.status_bar.config(text=f"开始自动分析，共 {len(all_files)} 个文件")

        # 创建后台线程
        self.auto_analysis_cancelled = False  # 修正初始化状态
        self.auto_analysis_thread = threading.Thread(
            target=self._auto_analysis_worker,
            args=(all_files,),
            daemon=True
        )
        self.auto_analysis_thread.start()
        self.root.after(100, self._handle_events)
        return True

    def analyze_code_chunk(self, chunk_info, file_path):
        chunk, line_start, line_end, chunk_type = chunk_info
        try:
            context_info = f"# 文件: {file_path.name}\n# 代码块类型: {chunk_type}\n# 行范围: {line_start}-{line_end}\n\n"
            chunk_with_context = context_info + chunk
            response = self.call_deepseek_api(chunk_with_context, file_path.suffix, file_path)

            # 特殊处理API验证错误
            if response['status_code'] == 401:
                # 如果是API验证错误，设置标志并通过事件队列通知主线程
                with self.validation_lock:
                    if not self.api_validation_error_shown:
                        self.api_validation_error_shown = True
                        self.api_validated = False
                        # 在API验证失败时，直接通知主线程更新状态栏
                        self.root.after(0, lambda: self.status_bar.config(text="API验证失败"))
                        # 返回False表示API验证失败
                        return False

            if response['status_code'] == 200:
                chunk_vulnerabilities = self.parse_response(response['text'], chunk.splitlines())
                for vuln in chunk_vulnerabilities:
                    vuln["行号"] = [line_start + line - 1 for line in vuln["行号"]]
                    vuln["文件路径"] = str(file_path)
                self.display_results(file_path, chunk_vulnerabilities)
                return True
            return None
        except Exception as e:
            self.log_error(f"代码块分析失败: {str(e)}")
            return None

    def _auto_analysis_worker(self, file_list):
        """自动分析的后台线程"""
        try:
            from concurrent.futures import ThreadPoolExecutor, as_completed

            # 在实际分析前验证API密钥有效性
            if not hasattr(self, 'api_validated') or not self.api_validated:
                # 显示验证中的状态
                self.root.after(0, lambda: self.status_bar.config(text="正在验证API密钥..."))
                self.root.after(0, lambda: self.root.config(cursor="wait"))  # 更改鼠标指针为等待状态

                try:
                    # 直接验证API密钥，避免异步验证和循环等待
                    result = self._validate_api_key(force_validation=True)

                    # 恢复鼠标指针
                    self.root.after(0, lambda: self.root.config(cursor=""))

                    if not result:
                        # 验证失败，无法继续分析
                        self.root.after(0, lambda: self.status_bar.config(text="API密钥无效，无法进行分析"))
                        self.root.after(0, lambda: messagebox.showerror("API验证失败",
                                                                        "API密钥验证失败，请检查API密钥是否正确。"))
                        self.root.after(0, lambda: self.btn_auto_analyze.config(text="自动分析"))
                        self.root.after(0, lambda: self.btn_analyze.config(state=tk.NORMAL))
                        return
                except Exception as e:
                    # 处理验证过程中的异常
                    self.root.after(0, lambda: self.root.config(cursor=""))
                    self.root.after(0, lambda: self.status_bar.config(text="API验证出错"))
                    self.root.after(0, lambda: messagebox.showerror("API验证错误", f"验证过程中发生错误：{str(e)}"))
                    self.root.after(0, lambda: self.btn_auto_analyze.config(text="自动分析"))
                    self.root.after(0, lambda: self.btn_analyze.config(state=tk.NORMAL))
                    return

            # 预处理阶段：计算总代码块数
            total_chunks = 0
            all_chunks = []
            valid_files = []

            # 先过滤出有效的文件
            for file_path in file_list:
                if file_path.suffix in self.supported_langs:
                    valid_files.append(file_path)

            # 只处理实际要分析的文件
            for file_path in valid_files:
                try:
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        code = f.read()
                        chunks = self._smart_code_chunking(code, file_path.suffix)
                        # 记录实际的分块
                        file_chunks = [(chunk, file_path) for chunk in chunks]
                        all_chunks.extend(file_chunks)
                        total_chunks += len(chunks)
                        print(f"[DEBUG] 文件 {file_path.name} 分为 {len(chunks)} 个代码块")
                except Exception as e:
                    self.log_error(f"读取文件失败: {str(e)}", file_path)
                    # 出错时不计入总块数

            # 设置进度条最大值
            self.root.after(0, lambda: self.progress.configure(maximum=total_chunks))
            self.status_bar.config(text=f"准备分析 {total_chunks} 个代码块")

            # 使用线程池处理
            processed_chunks = 0
            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = []
                for chunk, file_path in all_chunks:
                    if self.auto_analysis_cancelled:
                        break
                    futures.append(executor.submit(self.analyze_code_chunk, chunk, file_path))

                for future in as_completed(futures):
                    if self.auto_analysis_cancelled:
                        break
                    try:
                        result = future.result()
                        # 检查API错误
                        if result is False and not self.api_validation_error_shown:
                            with self.validation_lock:
                                if not self.api_validation_error_shown:
                                    self.api_validation_error_shown = True
                                    self.api_validated = False
                                    self.root.after(0, lambda: self.status_bar.config(text="API验证失败"))
                                    self.root.after(0, lambda: messagebox.showerror("API验证失败",
                                                                                    "API密钥验证失败，请检查API密钥是否正确。"))
                                    self.root.after(0, lambda: self.btn_auto_analyze.config(text="自动分析"))
                                    self.root.after(0, lambda: self.btn_analyze.config(state=tk.NORMAL))
                                    break
                    except Exception as e:
                        self.log_error(f"处理分析结果时出错: {str(e)}")
                    processed_chunks += 1
                    self.event_queue.put(('progress', processed_chunks, None))

        finally:
            # 使用root.after确保在主线程中安排事件处理
            if self.api_validation_error_shown:
                # 使用lambda避免直接调用方法，确保在UI线程中执行
                self.root.after(0, lambda: self.event_queue.put(('done', None, None)))
            else:
                # 非API验证失败情况下，正常发送done事件
                self.root.after(0, lambda: self.event_queue.put(('done', None, None)))

            # 直接在UI线程中重新启用按钮，确保按钮状态正确恢复
            self.root.after(0, lambda: self.btn_auto_analyze.config(text="自动分析", state=tk.NORMAL))
            self.root.after(0, lambda: self.btn_analyze.config(text="开始分析", state=tk.NORMAL))

            self._reset_analysis_state()

    def _init_file_tree(self, parent):
        """初始化文件树组件（支持拖动调整宽度）"""
        # 创建左侧面板
        self.left_panel = ttk.Frame(parent)
        parent.add(self.left_panel, weight=1)  # 添加到PanedWindow，weight=1表示拖动时会调整大小

        # 创建滚动容器
        container = ttk.Frame(self.left_panel)
        container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 树状视图 - 移除垂直滚动条
        self.tree = ttk.Treeview(
            container,
            selectmode='browse',
            show='tree',
            height=20
        )
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 动态列宽配置
        self.tree.column("#0", minwidth=200, stretch=tk.YES)

        # 绑定目录展开事件
        self.tree.bind("<<TreeviewOpen>>", self._handle_tree_resize)
        self.tree.bind("<<TreeviewClose>>", self._handle_tree_resize)
        self.tree.bind("<Configure>", self._handle_tree_resize)

    def _handle_tree_resize(self, event=None):
        """处理树状视图尺寸变化"""
        try:
            tree_width = self.tree.winfo_width()
            if tree_width < 150:
                tree_width = 150
            self.tree.column('#0', width=int(tree_width * 0.95))
        except Exception as e:
            self.log_error(f"树大小调整错误: {str(e)}")

    def _on_tree_open(self, event):
        """处理树节点展开事件"""
        item = self.tree.focus()
        if item:
            self.tree.item(item, open=True)

    def _on_tree_close(self, event):
        """处理树节点关闭事件"""
        item = self.tree.focus()
        if item:
            self.tree.item(item, open=False)

    def _init_toolbar(self, parent):
        """初始化工具栏"""
        self.toolbar = ttk.Frame(parent)
        self.toolbar.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)

        self.btn_api_settings = ttk.Button(self.toolbar, text="API设置", command=self.open_api_settings)
        self.btn_api_settings.pack(side=tk.LEFT, padx=5)

        ttk.Button(self.toolbar, text="打开项目", command=self.open_project).pack(side=tk.LEFT, padx=5)

        # 模型选择下拉框
        self.model_var = tk.StringVar(value="deepseek-chat")
        self.model_combobox = ttk.Combobox(self.toolbar, textvariable=self.model_var,
                                           values=["deepseek-chat", "deepseek-reasoner"],
                                           state="readonly", width=15)
        self.model_combobox.pack(side=tk.LEFT, padx=5)

        self.btn_analyze = ttk.Button(self.toolbar, text="开始分析", command=self.toggle_analysis)
        self.btn_analyze.pack(side=tk.LEFT, padx=5)

        # 自动分析按钮
        self.btn_auto_analyze = ttk.Button(self.toolbar, text="自动分析", command=self.toggle_auto_analysis)
        self.btn_auto_analyze.pack(side=tk.LEFT, padx=2)

        # 导出漏洞按钮
        self.btn_export = ttk.Button(self.toolbar, text="导出漏洞", command=self.export_vulnerabilities)
        self.btn_export.pack(side=tk.LEFT, padx=2)

    def toggle_analysis(self):
        """切换开始/取消分析状态"""
        if self.btn_analyze.cget("text") == "开始分析":
            # 获取选中的文件列表，检查是否有选中文件
            file_list = self._get_selected_files()
            if not file_list:
                messagebox.showinfo("提示", "请先选择要分析的文件")
                return

            # 切换到分析状态
            self.btn_analyze.config(text="取消分析")
            self.btn_auto_analyze.config(state=tk.DISABLED)
            # 调用start_analysis并检查其返回值
            result = self.start_analysis()
            # 如果start_analysis返回False（例如API密钥未配置），恢复按钮状态
            if result is False:
                self.btn_analyze.config(text="开始分析")
                self.btn_auto_analyze.config(state=tk.NORMAL)
        else:
            # 取消分析
            self.cancel_analysis()
            # 确保自动分析按钮被重新启用
            self.btn_auto_analyze.config(state=tk.NORMAL)

    def toggle_auto_analysis(self):
        """切换自动分析状态"""
        try:
            if self.btn_auto_analyze.cget("text") == "自动分析":
                # 开始新的分析
                self.api_validation_error_shown = False  # 重置API错误标志
                self.btn_auto_analyze.config(text="取消分析")
                self.auto_analyze()
            else:
                # 取消当前分析
                self.auto_analysis_cancelled = True
                # 立即恢复按钮状态，不等待事件处理
                self.btn_auto_analyze.config(text="自动分析", state=tk.NORMAL)
                self.btn_analyze.config(text="开始分析", state=tk.NORMAL)
        except Exception as e:
            self.log_error(f"切换分析状态失败: {str(e)}")
            self.btn_auto_analyze.config(text="自动分析")  # 确保恢复按钮状态

    def cancel_analysis(self):
        """取消分析操作"""
        # 先设置取消标志
        self.auto_analysis_cancelled = True
        self.status_bar.config(text="正在取消分析...")

        # 立即恢复按钮状态，不等待事件处理
        self.btn_analyze.config(text="开始分析", state=tk.NORMAL)
        self.btn_auto_analyze.config(text="自动分析", state=tk.NORMAL)

        # 手动添加done事件，确保分析状态被正确重置
        self.root.after(0, lambda: self.event_queue.put(('done', None, None)))

    def retry_api_validation(self):
        """重试API验证"""
        # 重置验证状态
        self.api_validated = False
        self.api_validation_error_shown = False

        # 显示验证中的状态
        self.status_bar.config(text="正在重新验证API密钥...")
        self.root.config(cursor="wait")  # 更改鼠标指针为等待状态

        # 启动异步验证
        self._validate_api_key(force_validation=True)

        # 验证超时计时
        validation_start_time = time.time()

        # 检查验证结果的函数
        def check_validation():
            try:
                # 尝试从事件队列获取验证结果
                event_type, result, _ = self.event_queue.get_nowait()

                if event_type == 'api_validation':
                    self.root.config(cursor="")
                    if result:
                        self.api_validated = True
                        self.status_bar.config(text="API密钥验证成功，可以开始分析...")
                        # 恢复按钮状态
                        self.btn_auto_analyze.config(text="自动分析", state=tk.NORMAL)
                        self.btn_analyze.config(state=tk.NORMAL)
                    else:
                        messagebox.showerror("API验证失败", "API密钥验证失败，请检查API密钥是否正确。")
                        self.status_bar.config(text="API验证失败")
                        self.api_validated = False
                    return
            except queue.Empty:
                # 检查是否超时
                current_time = time.time()
                if current_time - validation_start_time > 15:  # 增加超时时间到15秒
                    self.root.config(cursor="")
                    self.status_bar.config(text="API连接超时，请检查网络")
                    messagebox.showwarning("API验证超时", "API服务器响应超时，请检查网络连接。")
                    # 恢复按钮状态
                    self.btn_auto_analyze.config(text="自动分析", state=tk.NORMAL)
                    self.btn_analyze.config(state=tk.NORMAL)
                    return

                # 继续等待验证结果
                self.root.after(20, check_validation)  # 缩短检查间隔到20ms

        # 开始检查验证结果
        self.root.after(20, check_validation)

        # 重置进度条
        if hasattr(self, 'progress_var'):
            self.progress_var.set(0)
        if hasattr(self, 'progress'):
            self.progress['value'] = 0

        # 更新漏洞列表视图，确保已扫描出的漏洞能显示在界面上
        self.update_vulnerability_treeview()

    def _reset_analysis_state(self):
        """重置分析状态"""
        self.auto_analysis_paused = False
        self.auto_analysis_cancelled = False

        # 按钮状态只在需要时设置，避免和其他地方设置的冲突
        if hasattr(self, 'btn_analyze') and self.btn_analyze.cget("text") != "开始分析":
            self.btn_analyze.config(text="开始分析", state=tk.NORMAL)
        if hasattr(self, 'btn_auto_analyze') and (self.btn_auto_analyze.cget("text") != "自动分析" or
                                                  self.btn_auto_analyze.cget("state") != tk.NORMAL):
            self.btn_auto_analyze.config(text="自动分析", state=tk.NORMAL)

    def _init_right_panel(self, parent):
        """初始化右侧面板（修改为使用PanedWindow）"""
        # 创建右侧面板
        self.right_panel = ttk.Frame(parent)
        parent.add(self.right_panel, weight=3)  # 添加到PanedWindow，weight=3表示右侧面板初始宽度是左侧的3倍

        # 创建垂直方向的PanedWindow，用于上下调整各个面板
        self.right_paned = ttk.PanedWindow(self.right_panel, orient=tk.VERTICAL)
        self.right_paned.pack(fill=tk.BOTH, expand=True)

        # 结果表格（放入垂直PanedWindow）
        self.result_frame = ttk.Frame(self.right_paned)
        self.right_paned.add(self.result_frame, weight=1)
        self._init_result_tree(self.result_frame)

        # 代码预览（放入垂直PanedWindow）
        self.code_frame = ttk.Frame(self.right_paned)
        self.right_paned.add(self.code_frame, weight=2)
        self._init_code_preview(self.code_frame)

        # 详细信息面板（放入垂直PanedWindow）
        self.detail_frame = ttk.Frame(self.right_paned)
        self.right_paned.add(self.detail_frame, weight=1)
        self._init_detail_panel(self.detail_frame)

        # 设置初始分隔位置
        self.root.after(200, self._set_initial_right_pane_position)

    def _set_initial_right_pane_position(self):
        """设置右侧垂直PanedWindow的初始分隔条位置"""
        try:
            # 确保窗口已经渲染
            self.root.update_idletasks()

            # 获取窗口高度
            window_height = self.right_panel.winfo_height()

            if window_height > 100:
                # 设置第一个分隔条位置（结果表格和代码预览之间）
                self.right_paned.sashpos(0, int(window_height * 0.3))

                # 设置第二个分隔条位置（代码预览和详细信息之间）
                self.right_paned.sashpos(1, int(window_height * 0.7))
            else:
                # 默认位置
                self.right_paned.sashpos(0, 200)
                self.right_paned.sashpos(1, 500)
        except Exception as e:
            self.log_error(f"设置右侧分隔条位置失败: {str(e)}")

    def _init_result_tree(self, parent=None):
        """初始化结果表格（确保列顺序正确）"""
        # 如果没有提供父容器，使用默认的right_panel
        if parent is None:
            parent = self.right_panel

        # 创建结果框架
        result_frame = ttk.Frame(parent)
        result_frame.pack(expand=True, fill=tk.BOTH)

        # 先创建Treeview组件
        self.result_tree = ttk.Treeview(
            result_frame,
            columns=('ID', '漏洞类型', '风险等级', '文件路径', '漏洞描述', '风险点', 'Payload', '修复建议'),
            show='headings',
            selectmode='extended'
        )

        # 配置列标题和宽度
        columns_config = {
            'ID': {'width': 25, 'anchor': 'center'},
            '漏洞类型': {'width': 100, 'anchor': 'center'},
            '风险等级': {'width': 100, 'anchor': 'center'},
            '文件路径': {'width': 200, 'anchor': 'center'},
            '漏洞描述': {'width': 200},
            '风险点': {'width': 200},
            'Payload': {'width': 150},
            '修复建议': {'width': 150}
        }

        # 设置行高，增加行间距
        style = ttk.Style()
        style.configure("Treeview", rowheight=29)  # 增加默认行高

        # 为不同风险等级设置更明显的样式
        if hasattr(self, 'severity_colors'):
            for severity, color in self.severity_colors.items():
                # 创建带有边框和更明显间距的标签样式
                self.result_tree.tag_configure(severity, background=color, font=('微软雅黑', 10, 'bold'))
        for col, config in columns_config.items():
            self.result_tree.heading(col, text=col)
            self.result_tree.column(col, width=config['width'], anchor=config.get('anchor', 'w'))

        # 修改打包顺序，先打包树形视图，再打包滚动条
        self.result_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 添加垂直滚动条
        vsb = ttk.Scrollbar(result_frame, orient="vertical", command=self.result_tree.yview)
        self.result_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # 创建右键菜单
        self.result_menu = tk.Menu(self.root, tearoff=0)
        self.result_menu.add_command(label="删除选中项", command=self.delete_selected_vulnerabilities)

        # 绑定右键菜单事件
        self.result_tree.bind("<Button-3>", self.show_context_menu)
        # 绑定Ctrl+A快捷键
        self.result_tree.bind("<Control-a>", self.select_all_vulnerabilities)

        # 最后添加事件绑定
        self.result_tree.bind('<Double-1>', self.jump_to_file)

    def show_context_menu(self, event):
        """显示右键菜单"""
        # 获取当前选中项
        selected_items = self.result_tree.selection()
        if selected_items:  # 只有在有选中项时才显示菜单
            self.result_menu.post(event.x_root, event.y_root)

    def select_all_vulnerabilities(self, event):
        """选中所有漏洞"""
        self.result_tree.selection_set(self.result_tree.get_children())
        return "break"  # 阻止默认的Ctrl+A行为

    def delete_selected_vulnerabilities(self):
        """删除选中的漏洞"""
        selected_items = self.result_tree.selection()
        if not selected_items:
            return

        # 确认删除
        if messagebox.askyesno("确认删除", "确定要删除选中的漏洞吗？"):
            # 获取所有选中项的文件路径和漏洞信息
            for item in selected_items:
                values = self.result_tree.item(item)['values']
                file_path = values[3]  # 文件路径在第4列

                # 从self.vulnerabilities中删除对应的漏洞
                if file_path in self.vulnerabilities:
                    # 根据其他信息(如漏洞类型、行号等)匹配并删除具体漏洞
                    self.vulnerabilities[file_path] = [
                        v for v in self.vulnerabilities[file_path]
                        if not (v["漏洞类型"] == values[1] and
                                v["风险等级"] == values[2] and
                                str(v["行号"]) == values[8])  # 使用所有可用的匹配条件
                    ]

                    # 如果该文件没有漏洞了，删除该文件的记录
                    if not self.vulnerabilities[file_path]:
                        del self.vulnerabilities[file_path]

                # 从树形视图中删除
                self.result_tree.delete(item)

            # 更新状态栏
            total_vulns = sum(len(vulns) for vulns in self.vulnerabilities.values())
            self.status_bar.config(text=f"共发现 {total_vulns} 个漏洞")

    def load_full_config(self):
        """后台加载完整配置"""
        if self.config_path.exists():
            self.config.read(self.config_path, encoding='utf-8')
            self.api_key = self.config.get('DEFAULT', 'API_KEY', fallback='')
            self.api_endpoint = self.config.get('DEFAULT', 'API_ENDPOINT')

    def delayed_init(self):
        """延迟初始化非关键组件"""
        if not self._validate_api_key():
            pass

    def jump_to_file(self, event):
        """双击结果跳转至对应文件（修复路径匹配问题）"""
        item = self.result_tree.identify_row(event.y)
        if not item:
            return

        # 获取文件路径（第4列，索引为3）并转换为绝对路径
        selected_path = Path(self.result_tree.item(item, 'values')[3]).resolve()

        # 在文件树中查找对应路径（新增调试日志）
        def search_tree(node):
            for child in self.tree.get_children(node):
                child_path = Path(self.tree.item(child, 'values')[0]).resolve()

                if child_path == selected_path:
                    # 展开父节点并滚动到可视区域
                    parent = self.tree.parent(child)
                    self.tree.item(parent, open=True)
                    self.tree.see(child)
                    self.tree.selection_set(child)
                    self.preview_code()  # 强制刷新预览
                    return True
                if search_tree(child):
                    return True
            return False

        # 从根节点开始搜索（新增搜索失败提示）
        if not search_tree(''):
            messagebox.showinfo("路径未找到", f"未在项目中找到文件：{selected_path.name}")

    def _init_code_preview(self, parent=None):
        """初始化代码预览区域"""
        # 如果没有提供父容器，使用默认的right_panel
        if parent is None:
            parent = self.right_panel

        # 新增搜索工具栏
        self.search_frame = ttk.Frame(parent)
        self.search_frame.pack(side=tk.TOP, fill=tk.X, pady=2)

        # 搜索输入框和按钮
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(self.search_frame, textvariable=self.search_var)
        self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        ttk.Button(self.search_frame, text="搜索", command=self.search_in_code).pack(side=tk.LEFT, padx=5)
        ttk.Button(self.search_frame, text="下一个", command=lambda: self.search_in_code(direction="next")).pack(
            side=tk.LEFT, padx=5)
        ttk.Button(self.search_frame, text="上一个", command=lambda: self.search_in_code(direction="prev")).pack(
            side=tk.LEFT, padx=5)

        # 代码显示区域
        code_container = ttk.Frame(self.code_frame)
        code_container.pack(fill=tk.BOTH, expand=True)

        # 行号显示
        self.line_number = tk.Text(code_container, width=4, padx=3, takefocus=0, border=0,
                                   background='#f0f0f0', state='disabled')
        self.line_number.pack(side=tk.LEFT, fill=tk.Y)

        # 代码文本区域
        self.code_text = tk.Text(code_container, wrap=tk.NONE, state='normal')
        self.code_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 垂直滚动条
        code_vsb = ttk.Scrollbar(code_container, orient="vertical", command=self.code_text.yview)

        # 自定义yscrollcommand函数，确保滚动条操作时同步更新行号
        def custom_yscrollcommand(first, last):
            code_vsb.set(first, last)
            # 确保行号与代码同步滚动
            if hasattr(self, 'line_number') and self.line_number:
                self.line_number.yview_moveto(first)
                self.update_line_numbers()

        self.code_text.configure(yscrollcommand=custom_yscrollcommand)
        code_vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # 水平滚动条
        code_hsb = ttk.Scrollbar(self.code_frame, orient="horizontal", command=self.code_text.xview)
        self.code_text.configure(xscrollcommand=code_hsb.set)
        code_hsb.pack(side=tk.BOTTOM, fill=tk.X)

        # 同步行号和代码滚动
        self.code_text.bind('<Configure>', self.update_line_numbers)
        self.code_text.bind("<MouseWheel>", self.sync_scroll)

        # 设置标签样式
        self.code_text.tag_configure('number', foreground='#0000FF')
        self.code_text.tag_configure('string', foreground='#008000')
        self.code_text.tag_configure('comment', foreground='#999999')
        self.code_text.tag_configure('keyword', foreground='#CC7832')
        self.code_text.tag_configure('operator', foreground='#000000')
        self.code_text.tag_configure('bracket', foreground='#000000')
        self.code_text.tag_configure('function', foreground='#006699')
        self.code_text.tag_configure('class', foreground='#007F7F')

        # 扩展颜色配置（新增更多语法元素）
        self.code_text.tag_configure('php_keyword', foreground='#CC7832', font='TkFixedFont 10 bold')
        self.code_text.tag_configure('java_keyword', foreground='#0033CC', font='TkFixedFont 10 bold')
        self.code_text.tag_configure('string', foreground='#008000')
        self.code_text.tag_configure('comment', foreground='#999999')
        self.code_text.tag_configure('number', foreground='#FF4500')  # 新增数字
        self.code_text.tag_configure('operator', foreground='#0000FF')  # 新增运算符
        self.code_text.tag_configure('bracket', foreground='#802080')  # 新增括号
        self.code_text.tag_configure('function', foreground='#006699')  # 新增函数调用
        self.code_text.tag_configure('classname', foreground='#007F7F')  # 新增类名

        # 新增HTML特定语法元素
        self.code_text.tag_configure('html_tag', foreground='#0000FF')
        self.code_text.tag_configure('html_tag_name', foreground='#0000CD', font='TkFixedFont 10 bold')
        self.code_text.tag_configure('html_attribute', foreground='#FF8C00')
        self.code_text.tag_configure('html_value', foreground='#008000')
        self.code_text.tag_configure('html_entity', foreground='#800080')
        self.code_text.tag_configure('html_doctype', foreground='#800080', font='TkFixedFont 10 italic')
        self.code_text.tag_configure('html_script', foreground='#A52A2A')
        self.code_text.tag_configure('html_style', foreground='#2E8B57')
        self.code_text.tag_configure('html_template', foreground='#CC7832')

        # 新增Java特定语法元素
        self.code_text.tag_configure('java_primitive', foreground='#0000FF', font='TkFixedFont 10')  # 基本类型
        self.code_text.tag_configure('java_annotation', foreground='#808000', font='TkFixedFont 10 italic')  # 注解
        self.code_text.tag_configure('java_constant', foreground='#660E7A', font='TkFixedFont 10 bold')  # 常量
        self.code_text.tag_configure('java_modifier', foreground='#7F0055', font='TkFixedFont 10 bold')  # 修饰符
        self.code_text.tag_configure('java_exception', foreground='#CC0000', font='TkFixedFont 10')  # 异常相关
        self.code_text.tag_configure('java_import', foreground='#008080', font='TkFixedFont 10')  # 导入语句

        # 新增PHP特定语法元素
        self.code_text.tag_configure('php_var', foreground='#9B59B6', font='TkFixedFont 10')  # PHP变量
        self.code_text.tag_configure('php_function', foreground='#006699', font='TkFixedFont 10 bold')  # PHP函数
        self.code_text.tag_configure('php_class', foreground='#007F7F', font='TkFixedFont 10 bold')  # PHP类
        self.code_text.tag_configure('php_namespace', foreground='#808000', font='TkFixedFont 10 italic')  # 命名空间
        self.code_text.tag_configure('php_superglobal', foreground='#CC0000', font='TkFixedFont 10 bold')  # 超全局变量
        self.code_text.tag_configure('php_constant', foreground='#660E7A', font='TkFixedFont 10 bold')

    def search_in_code(self, direction=None):
        """在代码预览区域中搜索文本

        Args:
            direction (str, optional): 搜索方向，可选值为 "next" 或 "prev"。默认为 None，表示从当前位置开始新搜索。
        """
        # 移除之前的高亮
        self.code_text.tag_remove('search_highlight', '1.0', tk.END)

        # 获取搜索文本
        search_text = self.search_var.get()
        if not search_text:
            return

        # 配置高亮样式（如果尚未配置）
        if not self.code_text.tag_cget('search_highlight', 'background'):
            self.code_text.tag_configure('search_highlight', background='#FFFF00', foreground='#000000')

        # 确定搜索起始位置
        if direction == "next":
            # 从当前位置之后开始搜索
            start_pos = self.code_text.index(tk.INSERT)
        elif direction == "prev":
            # 从当前位置之前开始搜索 - 修复上一个按钮功能
            start_pos = self.code_text.index(tk.INSERT)
        else:
            # 新搜索，从文本开头开始
            start_pos = '1.0'

        # 执行搜索
        if direction == "prev":
            # 向上搜索
            pos = self.code_text.search(search_text, start_pos, '1.0', backwards=True)
            if not pos:
                # 如果没找到，则从末尾继续搜索
                pos = self.code_text.search(search_text, tk.END, '1.0', backwards=True)
        else:
            # 向下搜索
            pos = self.code_text.search(search_text, start_pos, tk.END)
            if not pos and start_pos != '1.0':
                # 如果没找到且不是从头开始，则从头继续搜索
                pos = self.code_text.search(search_text, '1.0', tk.END)

        # 处理搜索结果
        if pos:
            # 计算匹配文本的结束位置
            end_pos = f"{pos}+{len(search_text)}c"

            # 高亮显示匹配文本
            self.code_text.tag_add('search_highlight', pos, end_pos)

            # 确保匹配文本可见
            self.code_text.see(pos)

            # 将插入点移动到匹配文本之前或之后，取决于搜索方向
            if direction == "prev":
                # 向上搜索时，将插入点移动到匹配文本之前
                self.code_text.mark_set(tk.INSERT, pos)
            else:
                # 向下搜索时，将插入点移动到匹配文本之后
                self.code_text.mark_set(tk.INSERT, end_pos)

            # 更新状态栏
            line = pos.split('.')[0]
            self.status_bar.config(text=f"找到匹配项，位于第 {line} 行")
        else:
            # 未找到匹配项
            self.status_bar.config(text=f"未找到 '{search_text}'")

    def _init_detail_panel(self, parent=None):
        """初始化详细信息面板"""
        # 如果没有提供父容器，使用默认的right_panel
        if parent is None:
            parent = self.right_panel

        # 创建详细信息文本框
        self.detail_text = tk.Text(
            parent,
            height=8,
            wrap=tk.WORD,
            state='normal'
        )
        self.detail_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 添加滚动条 - 注意这里修改了side为RIGHT
        detail_vsb = ttk.Scrollbar(parent, orient="vertical", command=self.detail_text.yview)
        self.detail_text.configure(yscrollcommand=detail_vsb.set)
        detail_vsb.pack(side=tk.RIGHT, fill=tk.Y)

    def _init_status_bar(self):
        """初始化状态栏"""
        self.progress = ttk.Progressbar(self.root, mode='determinate')
        self.progress.pack(side=tk.BOTTOM, fill=tk.X, padx=5, pady=2)
        self.status_bar = ttk.Label(self.root, text="就绪", relief=tk.SUNKEN)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def _bind_events(self):
        """绑定事件"""
        self.result_tree.bind('<<TreeviewSelect>>', self.show_detail)
        self.tree.bind('<Double-1>', self.preview_code)
        self.code_text.bind('<MouseWheel>', self.sync_scroll)
        self.code_text.bind('<Configure>', self.update_line_numbers)
        self.root.bind('<Control-f>', lambda e: self.search_entry.focus())

    def delayed_config_init(self):
        """延迟配置初始化，确保UI组件已完全加载"""
        self.project_path = Path.cwd()
        threading.Thread(target=self._config_worker, daemon=True).start()

    def _config_worker(self):
        """配置初始化工作线程"""
        self._init_configuration()
        self.root.after(100, self._post_config_init)

    def _post_config_init(self):
        """配置初始化后的处理"""
        if not hasattr(self, 'api_key') or not self.api_key:
            self.api_key = self.config.get('DEFAULT', 'API_KEY', fallback='')

        self.validation_start_time = time.time()
        self._validate_api_key()
        self.root.after(500, self.check_initial_validation)

    def post_config_init(self):
        """配置加载后的初始化"""
        if not hasattr(self, 'api_key') or not self.api_key:
            self.api_key = self.config.get('DEFAULT', 'API_KEY', fallback='')
            return

        # 添加验证超时计时器
        self.validation_start_time = time.time()
        self.status_bar.config(text="正在验证API密钥...")
        self._validate_api_key()
        self.root.after(500, self.check_initial_validation)

    def check_initial_validation(self):
        """检查初始API验证状态 - 已移除实际验证逻辑"""
        self.api_validated = True
        return True

    def search_text(self, forward=True):
        """执行文本搜索"""
        self.code_text.tag_remove('highlight', '1.0', tk.END)
        search_str = self.search_entry.get()
        if not search_str:
            return

        start_pos = self.code_text.index(tk.INSERT)
        if forward:
            match_pos = self.code_text.search(search_str, start_pos, tk.END, nocase=False)
            if not match_pos:
                match_pos = self.code_text.search(search_str, '1.0', tk.END, nocase=False)
        else:
            match_pos = self.code_text.search(search_str, start_pos + '-1c', '1.0', backwards=True, nocase=False)
            if not match_pos:
                match_pos = self.code_text.search(search_str, tk.END, '1.0', backwards=True, nocase=False)

        if match_pos:
            end_pos = f"{match_pos}+{len(search_str)}c"
            self.code_text.tag_add('highlight', match_pos, end_pos)
            self.code_text.see(match_pos)
            self.code_text.mark_set(tk.INSERT, end_pos if forward else match_pos)

    # ------------------ 核心功能方法 ------------------ #
    def open_project(self):
        """打开项目目录"""
        if path := filedialog.askdirectory():
            self.project_path = Path(path)
            # 清空现有树节点
            self.tree.delete(*self.tree.get_children())
            # 添加根目录节点（新增代码）
            root_node = self.tree.insert('', 'end',
                                         text=self.project_path.name,
                                         values=[str(self.project_path)],
                                         open=True)
            # 填充子项（修改调用参数）
            self._populate_tree(self.project_path, parent=root_node)
            self.status_bar.config(text=f"已打开项目：{path}")

    def start_analysis(self):
        """开始分析（优化后的多线程版本）"""
        try:
            # 检查是否已有分析线程在运行
            if hasattr(self,
                       'auto_analysis_thread') and self.auto_analysis_thread and self.auto_analysis_thread.is_alive():
                messagebox.showinfo("提示", "已有分析任务正在运行，请等待完成或取消当前任务")
                return False

            # 获取选中的文件列表
            file_list = self._get_selected_files()

            # 检查是否有选中文件
            if not file_list:
                messagebox.showinfo("提示", "请先选择要分析的文件")
                return False

            # 检查API密钥是否已配置
            if not self.api_key:
                messagebox.showwarning("警告", "API密钥未配置，请先在设置中配置API密钥")
                return False

            # 重置API验证错误标志
            self.api_validation_error_shown = False

            # 优化API验证逻辑：只在实际需要使用API时才验证
            # 如果API已经验证过且有效，直接继续分析流程
            if hasattr(self, 'api_validated') and self.api_validated:
                # API已验证，直接继续
                pass
            else:
                # 设置API为已验证状态，跳过验证步骤
                # 只有在实际调用API进行分析时才会真正验证API有效性
                self.api_validated = True
                self.status_bar.config(text="准备开始分析...")

                # 注意：实际的API验证将在发送请求时进行
                # 这样可以避免不必要的API验证，提高用户体验

            # 记录开始分析
            self.log_info(f"开始分析任务，选中文件数: {len(file_list)}")

            # 初始化总块数计算 - 先不要设置进度条
            total_chunks = 0
            for file_path in file_list:
                if file_path.suffix in self.supported_langs:
                    # 对于大文件或特殊文件预计会有更多块
                    if file_path.suffix.lower() in ['.xml', '.pom', '.java',
                                                    '.php'] or file_path.name.lower() == 'pom.xml':
                        # 读取文件以估计块数
                        try:
                            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                                code = f.read()
                                chunks = self._smart_code_chunking(code, file_path.suffix)
                                total_chunks += len(chunks)
                        except Exception:
                            # 读取失败时默认为3个块
                            total_chunks += 3
                    else:
                        # 普通文件默认为1个块
                        total_chunks += 1

            # 设置进度条最大值为估计的总块数，而不是文件数
            total_chunks = max(total_chunks, len(file_list))
            self.progress['maximum'] = total_chunks
            self.progress['value'] = 0
            self.status_bar.config(text=f"准备分析 {len(file_list)} 个文件，预计 {total_chunks} 个代码块")

            # 重置分析状态
            self.auto_analysis_cancelled = False
            self.auto_analysis_paused = False

            # 清空事件队列，避免之前的事件影响当前分析
            try:
                while True:
                    self.event_queue.get_nowait()
                    self.event_queue.task_done()
            except queue.Empty:
                pass

            # 更改按钮文本和状态
            self.btn_analyze.config(text="取消分析", command=self.cancel_analysis)

            # 更新状态栏
            self.status_bar.config(text=f"准备分析 {len(file_list)} 个文件...")

            # 创建后台线程
            self.auto_analysis_thread = threading.Thread(
                target=self._analysis_worker,
                args=(file_list,),
                daemon=True,
                name="AnalysisWorkerThread"
            )
            self.auto_analysis_thread.start()
            self.root.after(100, self._handle_events)

            # 记录线程启动
            self.log_info(f"分析线程已启动，线程ID: {self.auto_analysis_thread.ident}")
            return True
        except Exception as e:
            self.log_error(f"启动分析失败: {str(e)}\n{traceback.format_exc()}")
            messagebox.showerror("错误", f"启动分析失败: {str(e)}")
            return False

    def _handle_events(self):
        """处理事件队列中的任务"""
        try:
            # 限制每次处理的事件数量，防止UI阻塞
            max_events = 10
            events_processed = 0

            while events_processed < max_events:
                try:
                    event_type, data, callback = self.event_queue.get_nowait()
                    events_processed += 1

                    # 如果分析已取消，忽略进度更新事件
                    if self.auto_analysis_cancelled and event_type == 'progress':
                        self.event_queue.task_done()
                        continue

                    # 根据事件类型处理
                    if event_type == 'progress':
                        self._handle_progress_event()
                    elif event_type == 'result':
                        self._handle_result_event(data)
                    elif event_type == 'done':
                        self._handle_done_event()
                        # 继续事件循环，不处理更多事件
                        self.root.after(100, self._handle_events)
                        return
                    elif event_type == 'api_validation':
                        self._handle_api_validation_event(data)
                    elif event_type == 'error':
                        self._handle_error_event(data)

                    # 执行回调函数（如果有）
                    if callback:
                        callback()

                    # 标记任务完成
                    self.event_queue.task_done()

                except queue.Empty:  # 队列为空时退出循环
                    break
                except Exception as e:
                    self.log_error(f"事件处理异常: {str(e)}")
                    continue
        except Exception as e:
            self.log_error(f"事件处理器异常: {str(e)}")
        finally:
            # 继续事件循环
            self.root.after(100, self._handle_events)

    def _handle_progress_event(self):
        """处理进度更新事件"""
        try:
            current_value = int(self.progress['value'])
            max_value = int(self.progress['maximum'])

            # 确保进度不超过最大值
            if current_value < max_value:
                self.progress.step(1)
                current_value += 1

            # 更新状态栏显示当前进度
            percentage = min(int((current_value / max_value) * 100) if max_value > 0 else 0, 100)
            # 使用"代码块"而不是"块"使表述更清晰
            self.status_bar.config(text=f"正在分析: {current_value}/{max_value} 代码块 ({percentage}%)")
        except Exception as e:
            self.log_error(f"更新进度失败: {str(e)}")

    def _handle_result_event(self, data):
        """处理结果事件"""
        try:
            file_path, vulnerabilities = data
            self._safe_display_results(file_path, vulnerabilities)
        except Exception as e:
            self.log_error(f"显示结果失败: {str(e)}")

    def _handle_done_event(self):
        """处理完成事件"""
        try:
            # 重置进度条和状态
            self.progress['value'] = 0

            # 恢复按钮状态，但只在必要时更改
            if hasattr(self, 'btn_analyze') and self.btn_analyze.cget("text") != "开始分析":
                self.btn_analyze.config(text="开始分析", command=self.start_analysis)

            if hasattr(self, 'btn_auto_analyze') and (self.btn_auto_analyze.cget("text") != "自动分析" or
                                                      self.btn_auto_analyze.cget("state") != tk.NORMAL):
                self.btn_auto_analyze.config(text="自动分析", state=tk.NORMAL)

            if hasattr(self, 'btn_export'):
                self.btn_export.config(state=tk.NORMAL)

            # 更新状态栏
            if hasattr(self, 'api_validation_error_shown') and self.api_validation_error_shown:
                self.status_bar.config(text="API验证失败")
            elif self.auto_analysis_cancelled:
                self.status_bar.config(text="分析已取消")
            else:
                # 检查是否有漏洞发现
                total_vulns = sum(len(vulns) for vulns in self.vulnerabilities.values())
                if total_vulns > 0:
                    self.status_bar.config(text=f"分析完成，共发现 {total_vulns} 个漏洞")
                else:
                    self.status_bar.config(text="分析完成，未发现漏洞")

            # 重置分析状态
            self.auto_analysis_cancelled = False
            self.auto_analysis_paused = False
            if hasattr(self, 'auto_analysis_thread'):
                delattr(self, 'auto_analysis_thread')
        except Exception as e:
            self.log_error(f"处理完成事件失败: {str(e)}")

    def _handle_api_validation_event(self, is_valid):
        """处理API验证事件"""
        try:
            if is_valid:
                self.status_bar.config(text="API密钥验证通过")
                self.api_validated = True
            else:
                self.status_bar.config(text="API密钥验证失败")
                self.api_validated = False
                if hasattr(self, 'btn_auto_analyze'):
                    self.btn_auto_analyze.config(text="自动分析", state=tk.NORMAL)
                if hasattr(self, 'btn_analyze'):
                    self.btn_analyze.config(state=tk.NORMAL)
                messagebox.showerror("API验证失败", "API密钥验证失败，请检查API密钥是否正确。")
        except Exception as e:
            self.log_error(f"处理API验证事件失败: {str(e)}")

    def _handle_error_event(self, error_msg):
        """处理错误事件"""
        try:
            self.show_error(error_msg)
            self.status_bar.config(
                text=f"发生错误: {error_msg[:50]}..." if len(error_msg) > 50 else f"发生错误: {error_msg}")
        except Exception as e:
            self.log_error(f"处理错误事件失败: {str(e)}")

    def _safe_display_results(self, file_path, vulnerabilities):
        """安全地在UI线程中显示结果"""
        try:
            # 初始化漏洞ID计数器（如果不存在）
            if not hasattr(self, 'vuln_id_counter'):
                self.vuln_id_counter = 1

            # 确保文件路径在漏洞字典中存在
            if file_path not in self.vulnerabilities:
                self.vulnerabilities[file_path] = []

            # 将新漏洞添加到现有列表
            self.vulnerabilities[file_path].extend(vulnerabilities)

            # 实时更新UI - 直接添加新发现的漏洞到树形视图
            for vuln in vulnerabilities:
                # 处理行号显示
                line_numbers = ", ".join(map(str, vuln["行号"])) if vuln["行号"] else "N/A"

                # 插入结果到Treeview
                try:
                    item_id = self.result_tree.insert(
                        '', 'end',
                        values=(
                            self.vuln_id_counter,
                            vuln["漏洞类型"],
                            vuln["风险等级"],
                            str(file_path),
                            vuln["详细描述"],
                            vuln.get("风险点", "N/A"),
                            vuln.get("Payload", "N/A"),
                            vuln.get("修复建议", "N/A"),
                            f"{line_numbers}"
                        )
                    )

                    # 设置行颜色（如果有定义）
                    if hasattr(self, 'severity_colors'):
                        severity = vuln["风险等级"]
                        color = self.severity_colors.get(severity, "#FFFFFF")
                        self.result_tree.tag_configure(severity, background=color)
                        self.result_tree.item(item_id, tags=(severity,))

                    # 递增漏洞ID计数器
                    self.vuln_id_counter += 1

                except KeyError as e:
                    self.log_error(f"字段缺失: {str(e)}")
                    continue

            # 更新状态栏显示总漏洞数
            total_vulns = sum(len(vulns) for vulns in self.vulnerabilities.values())
            self.status_bar.config(text=f"共发现 {total_vulns} 个漏洞")

        except Exception as e:
            self.log_error(f"显示结果失败: {str(e)}")

    def _analysis_worker(self, file_list):
        """优化的后台分析线程"""
        start_time = time.time()
        self.log_info(f"开始分析任务，文件数量: {len(file_list)}")

        try:
            # 计算总分块数量
            total_chunks = 0
            valid_files = []

            for file_path in file_list:
                if self.auto_analysis_cancelled:
                    self.log_info("分析已取消，停止预处理")
                    break

                # 检查文件是否存在
                if not file_path.exists():
                    self.log_error(f"文件不存在，跳过: {file_path}")
                    continue

                # 读取文件内容以估计分块数量
                try:
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        code = f.read()

                    # 估计分块数量
                    if len(code.splitlines()) > 1:
                        # 使用智能分块估计数量
                        chunks = self._smart_code_chunking(code, file_path.suffix)
                        chunk_count = len(chunks)
                        total_chunks += chunk_count
                        self.log_info(f"文件 {file_path.name} 预计分为 {chunk_count} 个代码块")
                    else:
                        # 小文件算作1个分块
                        total_chunks += 1

                    # 添加到有效文件列表
                    valid_files.append(file_path)
                except Exception as e:
                    # 读取失败时默认为1个分块
                    total_chunks += 1
                    valid_files.append(file_path)
                    self.log_error(f"估计分块数量失败: {str(e)}", file_path)

            # 检查是否有有效文件
            if not valid_files:
                self.log_info("没有有效文件可分析")
                self.root.after(0, lambda: self.status_bar.config(text="没有有效文件可分析"))
                return

            # 记录实际分块总数但不修改进度条最大值 (已在start_analysis中设置)
            self.log_info(f"总计划分为 {total_chunks} 个代码块")
            self.root.after(0, lambda: self.status_bar.config(text=f"开始分析 {total_chunks} 个代码块"))

            # 使用线程池管理线程
            import concurrent.futures
            max_workers = min(10, len(valid_files))  # 限制最大线程数
            self.log_info(f"启动分析线程池，最大线程数: {max_workers}")

            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                # 提交所有文件分析任务
                future_to_file = {executor.submit(self.analyze_file, file_path): file_path for file_path in valid_files}

                # 处理完成的任务
                for future in concurrent.futures.as_completed(future_to_file):
                    file_path = future_to_file[future]
                    try:
                        # 获取结果（如果有）
                        future.result()
                    except Exception as e:
                        self.log_error(f"文件 {file_path.name} 分析异常: {str(e)}")

        except Exception as e:
            self.log_error(f"分析工作线程异常: {str(e)}\n{traceback.format_exc()}")
        finally:
            # 记录总耗时
            elapsed_time = time.time() - start_time
            self.log_info(f"分析任务完成，总耗时: {elapsed_time:.2f}秒")

            # 只有在非API验证失败的情况下才发送done事件
            if not (hasattr(self, 'api_validation_error_shown') and self.api_validation_error_shown):
                self.root.after(0, lambda: self.event_queue.put(('done', None, None)))

            # 直接在UI线程中重新启用按钮，确保按钮状态正确恢复
            self.root.after(0, lambda: self.btn_auto_analyze.config(text="自动分析", state=tk.NORMAL))
            self.root.after(0, lambda: self.btn_analyze.config(text="开始分析", state=tk.NORMAL))

    def call_deepseek_api(self, code, suffix, file_path):
        """调用DeepSeek API"""
        # 前置校验
        if len(code.strip()) < 10:
            self.status_bar.config(text="代码内容过短或为空")
            return {'status_code': 400, 'text': '代码内容过短或为空'}
        if suffix not in self.supported_langs:
            self.status_bar.config(text=f"不支持的文件类型: {suffix}")
            return {'status_code': 400, 'text': '不支持的文件类型'}

        # 在实际调用API前验证API密钥有效性
        if not hasattr(self, 'api_validated') or not self.api_validated:
            self.status_bar.config(text="正在验证API密钥...")
            self.root.config(cursor="wait")  # 更改鼠标指针为等待状态

            # 简化API验证逻辑
            validation_result = self._validate_api_key(force_validation=True)
            self.root.config(cursor="")  # 恢复鼠标指针

            if not validation_result:
                self.status_bar.config(text="API密钥无效，无法进行分析")
                messagebox.showerror("API验证失败", "API密钥验证失败，请检查API密钥是否正确。")
                # 更改UI线程中的标志，确保事件处理程序知道API验证失败
                self.root.after(0, lambda: setattr(self, 'api_validation_error_shown', True))
                return {'status_code': 401, 'text': 'API密钥验证失败'}

        # 构建查询提示词
        prompt = f"""
        {self.config.get('DEFAULT', 'PROMPT_TEMPLATE', fallback='')}，没有漏洞就在漏洞类型处写无，严格按照以下JSON格式返回结果：
        {{
            "文件路径": "{str(file_path)}",
            "行号": [行号1, 行号2, ...],
            "风险等级": "高危/中危/低危",
            "漏洞类型": "代码执行/文件上传/XXE...",
            "详细描述": "漏洞具体描述",
            "风险点": "代码片段",
            "Payload": "实际攻击代码/输入示例",
            "修复建议": "修复建议"
        }}

        代码：
        {code}
        """

        # 创建带重试机制的Session
        session = requests.Session()
        retries = Retry(
            total=3,
            backoff_factor=1.0,
            status_forcelist=[500, 502, 503, 504],
            allowed_methods=["POST"]
        )
        session.mount('https://', HTTPAdapter(max_retries=retries))

        try:
            # 构建请求体JSON
            request_json = {
                "model": self.model_var.get(),
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.1,
                "max_tokens": 8192
            }

            # 发送请求
            response = session.post(
                self.api_endpoint,
                headers={"Authorization": f"Bearer {self.api_key}"},
                json=request_json,
                timeout=(10, 60)
            )

            # 处理API响应错误
            if response.status_code != 200:
                # 处理认证错误
                if response.status_code == 401:
                    error_message = "API密钥无效"
                    try:
                        response_json = json.loads(response.text)
                        if 'error' in response_json and 'message' in response_json['error']:
                            error_message = response_json['error']['message']
                    except:
                        pass

                    self.status_bar.config(text=f"API认证失败: {error_message}")
                    self.api_validated = False  # 重置验证状态

                return {
                    'status_code': response.status_code,
                    'text': response.text
                }

            return {
                'status_code': 200,
                'text': response.text
            }

        except requests.exceptions.Timeout as e:
            # 特殊处理超时错误，不显示弹窗
            error_msg = f"API请求超时: {str(e)}"
            self.log_error(error_msg, file_path)
            self.status_bar.config(text="API请求超时，请稍后重试")
            return {'status_code': 408, 'text': error_msg}  # 使用408状态码表示超时
        except requests.exceptions.ConnectionError as e:
            # 特殊处理连接错误，不显示弹窗
            error_msg = f"API连接失败: {str(e)}"
            self.log_error(error_msg, file_path)
            self.status_bar.config(text="API连接失败，请检查网络")
            return {'status_code': 503, 'text': error_msg}  # 使用503状态码表示服务不可用
        except requests.exceptions.RequestException as e:
            error_msg = f"API请求失败: {str(e)}"
            self.log_error(error_msg, file_path)
            # 不要在这里显示错误弹窗，而是返回错误信息
            return {'status_code': 500, 'text': error_msg}

    # ------------------ 辅助方法 ------------------ #
    def _save_config(self):
        """保存配置文件并立即应用更改"""
        with open(self.config_path, 'w', encoding='utf-8') as f:
            self.config.write(f)

        # 立即重新加载配置并应用更改
        self.reload_config()

        # 更新状态栏提示
        self.status_bar.config(text="配置已更新并应用")
        # 3秒后恢复状态栏
        self.root.after(3000, lambda: self.status_bar.config(text="就绪"))

    def reload_config(self):
        """重新加载配置文件并应用更改，无需重启程序"""
        # 重新读取配置文件
        if self.config_path.exists():
            self.config.read(self.config_path, encoding='utf-8')

            # 更新API相关配置
            self.api_key = self.config.get('DEFAULT', 'API_KEY', fallback='')
            self.api_endpoint = self.config.get('DEFAULT', 'API_ENDPOINT',
                                                fallback='https://api.deepseek.com/v1/chat/completions')

            # 更新其他可能的配置项
            timeout = self.config.get('DEFAULT', 'TIMEOUT', fallback='30')

            # 如果有主题设置，应用主题
            if 'THEME' in self.config['DEFAULT']:
                theme = self.config.get('DEFAULT', 'THEME', fallback='light')
                # 这里可以添加主题切换逻辑

            # 更新状态栏
            self.status_bar.config(text="配置已更新")

            # 3秒后恢复状态栏
            self.root.after(3000, lambda: self.status_bar.config(text="就绪"))

            # 重置API验证状态，以便在API密钥更改后重新验证
            if hasattr(self, 'api_validated'):
                self.api_validated = False

            return True
        return False

    def _validate_api_key(self, force_validation=False):
        """验证API密钥有效性并返回验证结果"""
        # 无API密钥则直接返回验证失败
        if not self.api_key:
            self.api_validated = False
            return False

        # 如果不是强制验证且已经验证过，直接返回缓存的结果
        if not force_validation and hasattr(self, 'api_validated'):
            return self.api_validated

        try:
            # 发送简单请求验证API密钥
            session = requests.Session()
            response = session.post(
                self.api_endpoint,
                headers={"Authorization": f"Bearer {self.api_key}"},
                json={
                    "model": "deepseek-coder",
                    "messages": [{"role": "user", "content": "验证API密钥"}],
                    "max_tokens": 10
                },
                timeout=10
            )

            # 根据响应状态码判断API密钥是否有效
            valid = (response.status_code == 200)
            self.api_validated = valid
            return valid

        except Exception as e:
            # 异常情况下视为验证失败
            self.log_error(f"API验证异常: {str(e)}")
            self.api_validated = False
            return False

    def open_api_settings(self):
        """打开API设置对话框"""
        # 创建一个顶层窗口
        settings_window = tk.Toplevel(self.root)
        settings_window.title("API设置")
        settings_window.geometry("700x450")  # 减小窗口尺寸，使其更加紧凑
        settings_window.resizable(True, True)  # 允许调整大小以适应不同屏幕分辨率
        settings_window.transient(self.root)
        settings_window.grab_set()  # 模态对话框

        # 内容框架 - 减小内边距使内容更加紧凑
        frame = ttk.Frame(settings_window, padding=10)
        frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # API密钥设置
        ttk.Label(frame, text="API密钥:").grid(row=0, column=0, sticky=tk.W, pady=5)
        api_key_var = tk.StringVar(value=self.api_key)
        api_key_entry = ttk.Entry(frame, textvariable=api_key_var, width=50)
        api_key_entry.grid(row=0, column=1, sticky=tk.W + tk.E, pady=5)

        # API终端设置
        ttk.Label(frame, text="API终端:").grid(row=1, column=0, sticky=tk.W, pady=5)
        api_endpoint_var = tk.StringVar(value=self.api_endpoint)
        api_endpoint_entry = ttk.Entry(frame, textvariable=api_endpoint_var, width=50)
        api_endpoint_entry.grid(row=1, column=1, sticky=tk.W + tk.E, pady=5)

        # 超时设置
        ttk.Label(frame, text="超时(秒):").grid(row=2, column=0, sticky=tk.W, pady=5)
        timeout_var = tk.StringVar(value=self.config.get('DEFAULT', 'TIMEOUT', fallback='30'))
        timeout_entry = ttk.Entry(frame, textvariable=timeout_var, width=10)
        timeout_entry.grid(row=2, column=1, sticky=tk.W, pady=5)

        # 提示模板设置
        ttk.Label(frame, text="提示模板:").grid(row=3, column=0, sticky=tk.W, pady=5)
        prompt_var = tk.StringVar(value=self.config.get('DEFAULT', 'PROMPT_TEMPLATE',
                                                        fallback='你是一个代码审计专家结合整段代码分析传参处理有没有可控点或者组件版本是否有漏洞，使用污点分析+AST分析对代码的语义信息进行安全分析'))
        prompt_entry = ttk.Entry(frame, textvariable=prompt_var, width=50)
        prompt_entry.grid(row=3, column=1, sticky=tk.W + tk.E, pady=5)

        # 余额显示区域
        balance_frame = ttk.LabelFrame(frame, text="API余额信息", padding=10)
        balance_frame.grid(row=4, column=0, columnspan=2, sticky=tk.W + tk.E, pady=10)

        # 余额信息标签
        balance_text = tk.Text(balance_frame, height=5, width=50, wrap=tk.WORD, state=tk.DISABLED)
        balance_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 查询余额函数
        def query_balance():
            # 禁用查询按钮，避免重复点击
            query_btn.config(state=tk.DISABLED)
            # 更新状态
            status_var.set("正在查询余额...")
            settings_window.update()

            # 使用当前输入框中的API密钥查询
            result = self.check_api_balance(api_key_var.get())

            # 启用查询按钮
            query_btn.config(state=tk.NORMAL)

            # 显示查询结果
            balance_text.config(state=tk.NORMAL)
            balance_text.delete(1.0, tk.END)

            if result["status"] == "success":
                try:
                    data = result["data"]
                    balance_info = ""

                    # 处理新格式的API响应 (balance_infos格式)
                    if "balance_infos" in data and data["balance_infos"]:
                        balance_info += f"账户状态: {'可用' if data.get('is_available', False) else '不可用'}\n"
                        balance_infos = data["balance_infos"]

                        for info in balance_infos:
                            currency = info.get("currency", "")
                            total = info.get("total_balance", "0")
                            granted = info.get("granted_balance", "0")
                            topped_up = info.get("topped_up_balance", "0")

                            balance_info += f"币种: {currency}\n"
                            balance_info += f"总余额: {total}\n"
                            balance_info += f"授予余额: {granted}\n"
                            balance_info += f"充值余额: {topped_up}\n\n"

                    # 处理旧格式的API响应 (grants格式)
                    elif "grants" in data and data["grants"]:
                        balance_info += f"账户: {data.get('title', 'Unknown')}\n\n"
                        grants = data["grants"]
                        for grant in grants:
                            name = grant.get("name", "Unknown")
                            amount = grant.get("amount", 0)
                            used = grant.get("used", 0)
                            available = grant.get("available", 0)
                            expires = grant.get("expires", "Unknown")

                            # 格式化金额显示
                            amount_str = f"${amount:.2f}" if isinstance(amount, (int, float)) else str(amount)
                            used_str = f"${used:.2f}" if isinstance(used, (int, float)) else str(used)
                            available_str = f"${available:.2f}" if isinstance(available, (int, float)) else str(
                                available)

                            balance_info += f"额度类型: {name}\n"
                            balance_info += f"总额: {amount_str}  已用: {used_str}  剩余: {available_str}\n"
                            if expires and expires != "Unknown":
                                # 尝试格式化过期时间
                                try:
                                    # 如果是ISO格式的时间字符串，转为更易读的格式
                                    if 'T' in expires and 'Z' in expires:
                                        dt = datetime.fromisoformat(expires.replace('Z', '+00:00'))
                                        expires = dt.strftime('%Y-%m-%d %H:%M:%S')
                                except:
                                    pass  # 如果转换失败，保持原格式
                                balance_info += f"过期时间: {expires}\n\n"

                    # 如果没有找到特定格式，则显示原始数据
                    else:
                        balance_info += "未识别的余额信息格式，显示原始数据:\n\n"
                        balance_info += str(data)

                    # 显示余额信息
                    balance_text.insert(tk.END, balance_info)
                    # 更新状态
                    status_var.set("余额查询成功")
                except Exception as ex:
                    # 处理解析数据异常
                    balance_text.insert(tk.END, f"解析余额信息失败: {str(ex)}\n")
                    if "data" in result:
                        balance_text.insert(tk.END, f"原始数据: {str(result['data'])}")
                    status_var.set("余额数据解析失败")
            else:
                # 显示错误信息
                error_message = result.get('message', '未知错误')
                balance_text.insert(tk.END, f"查询失败: {error_message}\n")

                # 如果有原始响应，也显示出来
                if 'raw' in result:
                    balance_text.insert(tk.END, f"原始响应: {result['raw'][:200]}...")

                # 更新状态
                status_var.set("余额查询失败")

            balance_text.config(state=tk.DISABLED)

        # 状态标签
        status_var = tk.StringVar(value="")
        status_label = ttk.Label(frame, textvariable=status_var, foreground="green")
        status_label.grid(row=5, column=0, sticky=tk.W, pady=10)

        # 按钮框架 - 移到状态标签同一行，靠右显示
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=5, column=1, pady=10, sticky=tk.E)

        # 应用配置按钮（应用后关闭窗口）
        def apply_settings():
            # 更新配置
            self.config['DEFAULT']['API_KEY'] = api_key_var.get()
            self.config['DEFAULT']['API_ENDPOINT'] = api_endpoint_var.get()
            self.config['DEFAULT']['TIMEOUT'] = timeout_var.get()
            self.config['DEFAULT']['PROMPT_TEMPLATE'] = prompt_var.get()

            # 保存配置并重新加载
            self._save_config()

            # 更新状态
            status_var.set("配置已应用")
            # 应用后关闭窗口
            settings_window.after(1000, settings_window.destroy)

        # 应用按钮 - 增加宽度
        apply_btn = ttk.Button(btn_frame, text="应用", command=apply_settings, width=20)
        apply_btn.pack(side=tk.LEFT, padx=25, pady=10)

        # 查询余额按钮 - 增加宽度
        query_btn = ttk.Button(btn_frame, text="查询余额", command=query_balance, width=20)
        query_btn.pack(side=tk.LEFT, padx=25, pady=10)

        # 居中显示窗口
        settings_window.update_idletasks()
        width = settings_window.winfo_width()
        height = settings_window.winfo_height()
        x = (settings_window.winfo_screenwidth() // 2) - (width // 2)
        y = (settings_window.winfo_screenheight() // 2) - (height // 2)
        settings_window.geometry(f"{width}x{height}+{x}+{y}")

        # 设置焦点到API密钥输入框
        api_key_entry.focus_set()

    def check_validation_result(self):
        """检查验证结果"""
        now = time.time()
        try:
            event_type, result, _ = self.event_queue.get_nowait()

            if event_type == 'api_validation':
                if result:
                    self.config['DEFAULT']['API_KEY'] = self.api_key
                    self._save_config()
                    self.api_validated = True  # 设置API验证状态为已验证成功
                    self.status_bar.config(text="API密钥验证通过")
                    return True
                else:
                    self.api_validated = False  # 设置API验证状态为验证失败
                    self.status_bar.config(text="API密钥无效")
                    return False
        except Empty:
            pass

        time_elapsed = now - self.validation_start_time

        if time_elapsed > 15:
            self.status_bar.config(text="API服务器响应超时")
            return

        self.root.after(50, self.check_validation_result)
        return None

    def retry_api_validation(self):
        """重试API验证 - 已移除实际验证逻辑"""
        self.api_validated = True
        return True

    def log_error(self, error_msg, file_path=None):
        """记录错误日志"""
        try:
            log_entry = f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {error_msg}"
            if file_path:
                log_entry += f" | 文件：{file_path}"

            with open(self.log_file, 'a', encoding='utf-8') as f:
                f.write(log_entry + '\n')
        except Exception as e:
            print(f"日志记录失败: {str(e)}")

    # ------------------ UI同步方法 ------------------ #
    def show_detail(self, event):
        """显示漏洞详情（带格式优化）"""
        selected = self.result_tree.selection()
        if not selected:
            return

        item = self.result_tree.item(selected[0])
        values = item['values']

        # 格式化修复建议（新增格式处理）
        repair_advice = values[7].replace('；', '\n')  # 中文分号转行
        repair_advice = re.sub(r'\d+\.\s*', '', repair_advice)  # 移除已有编号

        details = f"""文件路径：{values[3]}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
风险等级：{values[2]}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
漏洞类型：{values[1]}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
风险点：{values[5]}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Payload：{values[6]}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
详细描述：{values[4]}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
修复建议：\n"""

        # 添加带格式的修复建议（新增内容）
        self.detail_text.config(state='normal')
        self.detail_text.delete(1.0, tk.END)

        # 插入基础信息
        self.detail_text.insert(tk.END, details)

        # 格式化修复建议（新增高亮逻辑）
        for i, line in enumerate(repair_advice.split('\n'), 1):
            if line.strip():
                self.detail_text.insert(tk.END, f"{i}. ", 'repair_title')
                self.detail_text.insert(tk.END, f"{line.strip()}\n", 'repair_content')

        self.detail_text.config(state='disabled')

        # 添加标签配置（新增样式配置）
        self.detail_text.tag_configure('repair_title', foreground='#0078D7', font=('微软雅黑', 10, 'bold'))
        self.detail_text.tag_configure('repair_content', foreground='#333333', font=('微软雅黑', 10))

    def sync_scroll(self, event=None):
        """同步滚动行号和代码"""
        if hasattr(self, 'line_number') and self.line_number and self.code_text:
            self.line_number.yview_moveto(self.code_text.yview()[0])
            self.update_line_numbers()
        return "break"

    def update_line_numbers(self, event=None):
        """更新行号"""
        if hasattr(self, 'line_number') and self.line_number and self.code_text:
            # 添加递归保护标志
            if hasattr(self, '_updating_line_numbers') and self._updating_line_numbers:
                return

            self._updating_line_numbers = True
            try:
                # 获取可见区域的第一行和最后一行
                first_line = self.code_text.index("@0,0").split('.')[0]
                last_line = self.code_text.index(f"@0,{self.code_text.winfo_height()}").split('.')[0]

                # 计算行数
                first = int(float(first_line))
                last = int(float(last_line)) + 1

                # 清除现有行号
                self.line_number.config(state='normal')
                self.line_number.delete(1.0, tk.END)

                # 计算行号宽度 - 更精确地计算所需宽度
                max_line = int(last_line)
                width = len(str(max_line)) + 1  # 添加一点额外空间
                self.line_number.config(width=width)

                # 添加新行号 - 右对齐显示
                for i in range(first, last):
                    self.line_number.insert(tk.END, f"{i:>{width - 1}}\n")  # 使用右对齐格式

                self.line_number.config(state='disabled')

                # 同步滚动位置 - 使用原始的yview方法
                if hasattr(self, '_original_code_text_yview'):
                    yview_pos = self._original_code_text_yview()[0]
                else:
                    yview_pos = self.code_text.yview()[0]

                self.line_number.yview_moveto(yview_pos)
            finally:
                self._updating_line_numbers = False

    def preview_code(self, event=None):
        """预览选中文件的代码"""
        # 统一获取选中项逻辑，无论是通过事件触发还是直接调用
        item = self.tree.selection()[0] if self.tree.selection() else None

        if not item:
            return  # 提前返回避免空指针

        try:
            # 获取完整路径并验证文件类型
            file_path = Path(self.tree.item(item, 'values')[0])
            if not file_path.is_file():
                return

            # 读取文件内容
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()

            # 清空并更新代码显示
            self.code_text.config(state='normal')
            self.code_text.delete(1.0, tk.END)
            self.code_text.insert(1.0, content)

            # 应用语法高亮 - 防止递归调用
            if hasattr(self, '_highlighting_in_progress') and self._highlighting_in_progress:
                return

            self._highlighting_in_progress = True
            try:
                self._apply_syntax_highlighting(file_path.suffix)
            finally:
                self._highlighting_in_progress = False

            # 更新行号
            self.update_line_numbers()

        except RecursionError:
            self.log_error(f"预览代码失败: 递归深度超出限制")
            messagebox.showerror("预览失败", "文件过大或结构过于复杂，无法完成预览")
        except Exception as e:
            self.log_error(f"预览代码失败: {str(e)}")
            messagebox.showerror("预览失败", f"无法读取文件: {str(e)}")

    def _apply_syntax_highlighting(self, file_ext):
        """应用语法高亮"""
        # 清除现有标签
        for tag in self.code_text.tag_names():
            if tag != "sel":  # 保留选择标签
                self.code_text.tag_remove(tag, "1.0", tk.END)

        # 获取代码内容
        code = self.code_text.get("1.0", tk.END)

        # 根据文件类型应用高亮
        if file_ext in self.supported_langs:
            self._highlight_code(code, file_ext)

        # 更新行号
        self.update_line_numbers()

        # 绑定滚动事件，确保行号与代码同步滚动
        if not hasattr(self, '_scroll_binding_set'):
            # 绑定各种可能触发滚动的事件
            self.code_text.bind("<Configure>", self.update_line_numbers)
            self.code_text.bind("<KeyPress>", self.update_line_numbers)
            self.code_text.bind("<KeyRelease>", self.update_line_numbers)
            self.code_text.bind("<MouseWheel>", self.update_line_numbers)

            # 保存原始的yview方法，避免递归调用
            self._original_code_text_yview = self.code_text.yview

            # 重写Text组件的yview方法以触发自定义事件
            def custom_yview(*args):
                # 调用原始方法获取结果
                result = self._original_code_text_yview(*args)

                # 避免递归调用
                if not hasattr(self, '_updating_line_numbers') or not self._updating_line_numbers:
                    self.update_line_numbers()

                return result

            self.code_text.yview = custom_yview

            # 查找代码文本区域的滚动条
            code_scrollbar = None
            for child in self.code_text.master.winfo_children():
                if isinstance(child, ttk.Scrollbar) and child.cget('orient') == 'vertical':
                    code_scrollbar = child
                    break

            # 如果找到滚动条，绑定其set方法
            if code_scrollbar:
                # 不直接替换set方法，而是使用monkey patching技术
                original_set = code_scrollbar.set

                def custom_set_wrapper(first, last):
                    # 调用原始set方法
                    original_set(first, last)

                    # 避免递归调用
                    if not hasattr(self, '_updating_line_numbers') or not self._updating_line_numbers:
                        self.update_line_numbers()
                    return None

                # 替换为包装后的方法
                code_scrollbar.set = custom_set_wrapper

            self._scroll_binding_set = True

    def _highlight_code(self, code, file_ext):
        """执行语法高亮"""
        # 公共语法元素（优化正则表达式和颜色配置）
        number_pattern = r'\b\d+\.?\d*\b'
        operator_pattern = r'(\+|\-|\*|\/|\%|\=|\&|\||\<|\>|!|\^|~|\?|:)'
        bracket_pattern = r'[{}()\[\]]'
        common_patterns = [
            (number_pattern, 'number', '#6897BB'),  # 数字 - 蓝紫色
            (operator_pattern, 'operator', '#A9B7C6'),  # 运算符 - 浅灰色
            (bracket_pattern, 'bracket', '#A9B7C6')  # 括号 - 浅灰色
        ]
        # 初始化patterns并添加返回条件
        patterns = []

        # 高亮公共元素
        for pattern, tag, *args in common_patterns:
            color = args[0] if args else None
            if color:
                self.code_text.tag_configure(tag, foreground=color)
            for match in self._find_matches(pattern, code):
                self.code_text.tag_add(tag, match[0], match[1])

        # HTML特定规则
        if file_ext == '.html' or file_ext == '.htm':
            patterns = [
                (r'<\!DOCTYPE.*?>', 'html_doctype', '#800080'),  # DOCTYPE声明
                (r'<\!--.*?-->', 'comment', '#999999'),  # 注释
                (r'</?[a-zA-Z][a-zA-Z0-9]*(?:\s+[a-zA-Z][a-zA-Z0-9]*(?:=(?:"[^"]*"|\'[^\']*\'|[^\s>]*))?)*\s*/?>',
                 'html_tag', '#0000FF'),  # 标签
                # 修复属性名和属性值的正则表达式，移除不兼容的后向查找
                (r'<[^>]*\s([a-zA-Z][a-zA-Z0-9\-]*)=', 'html_attribute', '#FF8C00', 1),  # 属性名
                (r'=("[^"]*"|\'[^\']*\')', 'html_value', '#008000', 1),  # 属性值
                (r'&[a-zA-Z0-9]+;', 'html_entity', '#800080'),  # HTML实体
                (r'<script\b[^>]*>(.*?)</script>', 'html_script', '#A52A2A', re.DOTALL),  # script标签
                (r'<style\b[^>]*>(.*?)</style>', 'html_style', '#2E8B57', re.DOTALL),  # style标签
                (r'{{.*?}}|\{%.*?%\}', 'html_template', '#CC7832')  # 模板语法(如Angular, Vue等)
            ]

            # 添加HTML标签关键字高亮 - 修复标签名称的正则表达式
            html_tags = [
                'html', 'head', 'body', 'div', 'span', 'p', 'a', 'img', 'table', 'tr', 'td', 'th',
                'ul', 'ol', 'li', 'form', 'input', 'button', 'select', 'option', 'textarea',
                'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'header', 'footer', 'nav', 'section', 'article',
                'aside', 'main', 'canvas', 'audio', 'video', 'source', 'iframe', 'script', 'style',
                'link', 'meta', 'title', 'base', 'br', 'hr', 'pre', 'code', 'em', 'strong', 'i', 'b'
            ]
            # 使用捕获组替代后向查找
            tag_pattern = r'</?(' + '|'.join(html_tags) + r')[\s>]'
            patterns.append((tag_pattern, 'html_tag_name', '#0000CD', 1))

        # XML特定规则
        elif file_ext == '.xml':
            patterns = [
                (r'<\?xml.*?\?>', 'xml_declaration', '#007F7F'),  # XML声明
                (r'<!--.*?-->', 'comment', '#999999'),  # 注释
                (r'<[^>]+>', 'tag', '#CC6600'),  # 标签
                (r'"[^"]*"', 'string', '#008000'),  # 属性值
                (r'\bxmlns\b', 'keyword', '#CC7832'),  # 命名空间
                (r'\bversion\b|\bencoding\b', 'attribute', '#9B59B6')  # 属性
            ]

        # PHP特定规则（新增变量高亮）
        elif file_ext == '.php':
            # 扩展PHP高亮
            patterns = [
                (r'\$[a-zA-Z_\x7f-\xff][a-zA-Z0-9_\x7f-\xff]*', 'php_var', '#9B59B6'),  # 紫色变量
                (r'\$_(GET|POST|REQUEST|COOKIE|SESSION|SERVER|ENV|FILES)', 'php_superglobal', '#CC0000'),  # 超全局变量
                (r'(function)\s+([a-zA-Z_\x7f-\xff][a-zA-Z0-9_\x7f-\xff]*)', 'php_function', ('#006699', 2)),  # 函数声明
                (r'(class|interface|trait)\s+(\w+)', 'php_class', ('#007F7F', 2)),  # 类名
                (r'(namespace)\s+([a-zA-Z0-9_\\]+)', 'php_namespace', ('#808000', 2)),  # 命名空间
                (r'(use)\s+([a-zA-Z0-9_\\]+)', 'php_namespace', ('#808000', 2)),  # use语句
                (r'\b([A-Z][A-Z0-9_]*)\b', 'php_constant', '#660E7A'),  # 常量
                (r'//.*?$|/\*.*?\*/', 'comment', '#999999'),  # 注释
                (r'(".*?"|\'.*?\')', 'string', '#008000'),  # 字符串
                (r'<\?php|\?>', 'php_tag', '#8B008B')  # PHP标签
            ]

            # 新增PHP关键字
            php_keywords = [
                'echo', 'print', 'die', 'exit', 'isset', 'empty',
                'include', 'require', 'include_once', 'require_once',
                'array', 'list', 'foreach', 'as', 'do', 'declare',
                'if', 'else', 'elseif', 'switch', 'case', 'default',
                'while', 'for', 'break', 'continue', 'return',
                'global', 'static', 'final', 'abstract', 'private', 'protected',
                'public', 'var', 'const', 'clone', 'try', 'catch', 'throw',
                'finally', 'instanceof', 'insteadof', 'interface', 'implements',
                'extends', 'trait', 'yield', 'yield from', 'fn', 'match'
            ]
            patterns.append((r'\b(' + '|'.join(php_keywords) + r')\b', 'php_keyword', '#CC7832'))

            # 新增PHP内置函数
            php_builtin_functions = [
                'array_merge', 'count', 'strlen', 'substr', 'str_replace',
                'preg_match', 'preg_replace', 'explode', 'implode', 'trim',
                'file_get_contents', 'file_put_contents', 'fopen', 'fclose',
                'mysql_query', 'mysqli_query', 'PDO', 'json_encode', 'json_decode'
            ]
            patterns.append((r'\b(' + '|'.join(php_builtin_functions) + r')\b(?=\s*\()', 'php_function', '#006699'))

        # Java特定规则（增强高亮显示 - 使用更深的颜色）
        elif file_ext == '.java':
            patterns = [
                # 基础元素
                (r'@[a-zA-Z_$][a-zA-Z\d_$]*', 'java_annotation', '#FF8C00'),  # 注解 - 深橙色
                (r'\b([A-Z_][A-Z0-9_]+)\b', 'java_constant', '#9370DB'),  # 常量 - 深紫色
                (r'//.*?$|/\*.*?\*/', 'java_comment', '#5D6C79'),  # 注释 - 深灰色
                (r'\".*?\"', 'java_string', '#008B45'),  # 字符串 - 深绿色
                (r'\'.*?\'', 'java_char', '#008B45'),  # 字符 - 深绿色

                # 类型和声明
                (r'(class|interface|enum|record)\s+(\w+)', 'java_class_decl', ('#1E6262', 2)),  # 类声明 - 深青色
                (r'\b(void|int|byte|short|long|float|double|char|boolean)\b', 'java_primitive', '#0000CD'),
                # 基本类型 - 深蓝色
                (r'\b([A-Z][a-zA-Z0-9_]*)\b(?!\s*\()', 'java_class_ref', '#1E6262'),  # 类引用 - 深青色

                # 方法和函数
                (r'(public|private|protected|static)?\s+([a-zA-Z_$][a-zA-Z\d_$]*)\s*\(', 'java_method_decl',
                 ('#E8A317', 2)),  # 方法声明 - 深橙色
                (r'\b([a-zA-Z_$][a-zA-Z\d_$]*)\s*\(', 'java_method_call', '#4F4F4F'),  # 方法调用 - 深灰色

                # 泛型
                (r'<([A-Z][a-zA-Z0-9_]*(?:\s*,\s*[A-Z][a-zA-Z0-9_]*)*)>', 'java_generic', '#7B68EE'),  # 泛型 - 深紫色
                (r'<\?\s+(?:extends|super)\s+[A-Z][a-zA-Z0-9_]*>', 'java_wildcard', '#7B68EE'),  # 通配符 - 深紫色

                # Lambda和函数式
                (r'\([^()]*\)\s*->\s*[^;]+', 'java_lambda', '#D2691E'),  # Lambda表达式 - 深橙色
                (r'::', 'java_method_ref', '#D2691E'),  # 方法引用 - 深橙色

                # 异常处理
                (r'\b(try|catch|throw|throws|finally)\b', 'java_exception', '#B22222'),  # 异常处理关键字 - 深红色

                # 包和导入
                (r'\b(import|package)\b.*?;', 'java_import', '#696969'),  # 导入语句 - 深灰色

                # 注释标签
                (r'@(param|return|throws|author|version|since|see|deprecated)\b', 'java_doc_tag', '#2E8B57'),
                # JavaDoc标签 - 深绿色

                # 修饰符
                (
                    r'\b(public|private|protected|static|final|abstract|synchronized|volatile|transient|native|strictfp)\b',
                    'java_modifier', '#D2691E'),  # 修饰符 - 深橙色

                # 数字字面量
                (r'\b(\d+\.?\d*[fFlL]?)\b', 'java_number', '#4169E1'),  # 数字 - 深蓝色
                (r'\b(0x[0-9a-fA-F]+)\b', 'java_hex', '#4169E1'),  # 十六进制 - 深蓝色

                # 特殊语法
                (r'\?|:', 'java_ternary', '#D2691E'),  # 三元运算符 - 深橙色
                (r'\b(instanceof)\b', 'java_instanceof', '#D2691E'),  # instanceof - 深橙色
                (r'\b(new)\b', 'java_new', '#D2691E'),  # new关键字 - 深橙色
            ]

            # Java关键字
            java_keywords = [
                'if', 'else', 'switch', 'case', 'default', 'for', 'do', 'while', 'break', 'continue',
                'return', 'this', 'super', 'extends', 'implements', 'null', 'true', 'false',
                'assert', 'enum', 'var', 'yield', 'sealed', 'permits', 'non-sealed', 'record'
            ]
            patterns.append((r'\b(' + '|'.join(java_keywords) + r')\b', 'java_keyword', '#D2691E'))  # 关键字 - 深橙色

            # Java注解关键字
            java_annotations = [
                'Override', 'Deprecated', 'SuppressWarnings', 'FunctionalInterface',
                'SafeVarargs', 'Target', 'Retention', 'Documented', 'Inherited'
            ]
            patterns.append(
                (r'@(' + '|'.join(java_annotations) + r')\b', 'java_std_annotation', '#DAA520'))  # 标准注解 - 深黄色

        # 应用语言特定规则
        for pattern, tag, *args in patterns:
            # 提取颜色和分组参数
            color = None
            group = 0
            flags = 0

            if args:
                if isinstance(args[0], tuple):  # 处理带分组的颜色配置
                    color = args[0][0]
                    group = args[0][1] if len(args[0]) > 1 else 0
                elif isinstance(args[0], int) and args[0] < 100:  # 处理捕获组索引
                    group = args[0]
                elif isinstance(args[0], int):  # 处理正则标志
                    flags = args[0]
                else:  # 普通颜色配置
                    color = args[0]

                if len(args) > 1 and isinstance(args[1], int):
                    if args[1] < 100:  # 假设捕获组索引小于100
                        group = args[1]
                    else:  # 正则标志
                        flags = args[1]

            try:
                for match in self._find_matches(pattern, code, flags):
                    try:
                        if len(match) > 2 and group > 0:  # 使用捕获组
                            # 获取原始匹配文本的位置
                            full_start_pos = match[0]
                            full_end_pos = match[1]

                            # 计算捕获组在原始文本中的相对位置
                            full_start_idx = int(full_start_pos.split('+')[1].split('c')[0])

                            # 在原始文本中查找捕获组
                            match_text = code[full_start_idx - 1:full_start_idx - 1 + (
                                    int(full_end_pos.split('+')[1].split('c')[0]) - full_start_idx)]

                            # 使用正则表达式重新匹配以获取捕获组
                            m = re.search(pattern, match_text, flags=flags | re.MULTILINE)
                            if m and m.group(group):
                                # 计算捕获组的位置
                                group_start = m.start(group) - m.start(0)
                                group_length = len(m.group(group))

                                # 计算最终位置
                                start = f"1.0 + {full_start_idx + group_start - 1}c"
                                end = f"1.0 + {full_start_idx + group_start + group_length - 1}c"

                                if color:
                                    self.code_text.tag_configure(tag, foreground=color)
                                self.code_text.tag_add(tag, start, end)
                        else:
                            # 处理普通匹配
                            start = match[0]
                            end = match[1]

                            if color:
                                self.code_text.tag_configure(tag, foreground=color)
                            self.code_text.tag_add(tag, start, end)
                    except Exception as e:
                        print(f"应用标签错误: {tag} - {str(e)}")
            except Exception as e:
                print(f"处理正则表达式错误: {pattern} - {str(e)}")

    def _find_matches(self, pattern, text, group=0):
        """查找所有匹配项，返回位置列表"""
        try:
            # 添加最大迭代次数限制，防止无限递归
            max_iterations = 10000
            iteration_count = 0

            matches = []
            start = "1.0"

            while True:
                iteration_count += 1
                if iteration_count > max_iterations:
                    # 达到最大迭代次数，中断处理
                    self.log_error(f"语法高亮处理中断: 达到最大迭代次数 {max_iterations}")
                    break

                pos = self.code_text.search(pattern, start, tk.END, regexp=True)
                if not pos:
                    break

                line, col = map(int, pos.split('.'))
                end_line, end_col = line, col

                # 计算匹配文本的结束位置
                match_text = self.code_text.get(pos, f"{pos} lineend")
                match = re.search(pattern, match_text)
                if not match:
                    # 如果没有匹配到，移动到下一行继续
                    start = f"{line + 1}.0"
                    continue

                # 计算匹配文本的结束位置
                match_length = match.end(group) - match.start(group)
                end_pos = f"{line}.{col + match_length}"

                # 添加匹配结果
                matches.append((f"{line}.{col + match.start(group)}", f"{line}.{col + match.end(group)}"))

                # 更新下一次搜索的起始位置
                start = end_pos

            return matches
        except RecursionError:
            self.log_error("语法高亮递归错误")
            return []
        except Exception as e:
            self.log_error(f"查找匹配项错误: {str(e)}")
            return []

    def _populate_tree(self, path, parent=''):
        """递归填充文件树"""
        try:
            # 获取排序后的目录和文件列表（保持排序逻辑）
            dirs = sorted([d for d in path.iterdir()
                           if d.is_dir() and not d.name.startswith('.')],
                          key=lambda x: x.name.lower())

            files = sorted([f for f in path.iterdir()
                            if f.is_file() and not f.name.startswith('.')],
                           key=lambda x: x.name.lower())

            # 先处理目录再处理文件（保持目录优先）
            for item in dirs:
                node = self.tree.insert(parent, 'end',
                                        text=item.name,
                                        values=[str(item)],
                                        open=False)
                # 添加一个临时子节点，表示该目录可以展开
                self.tree.insert(node, 'end', text="加载中...", values=["loading"])

                # 绑定展开事件，只有当用户展开目录时才加载子目录内容
                if not hasattr(self, '_tree_expand_handler_set'):
                    self.tree.bind('<<TreeviewOpen>>', self._on_tree_expand)
                    self._tree_expand_handler_set = True

            # 处理文件
            for item in files:
                self.tree.insert(parent, 'end',
                                 text=item.name,
                                 values=[str(item)])
        except Exception as e:
            messagebox.showerror("目录错误", f"读取失败: {str(e)}")

    def _on_tree_expand(self, event):
        """当树节点展开时异步加载子节点"""
        item_id = self.tree.focus()

        # 检查是否已经加载过
        children = self.tree.get_children(item_id)
        if len(children) == 1 and self.tree.item(children[0], "values")[0] == "loading":
            # 删除加载中的临时节点
            self.tree.delete(children[0])

            # 获取目录路径
            item_path = Path(self.tree.item(item_id, "values")[0])

            # 使用after方法在主线程空闲时异步加载
            self.root.after(10, lambda: self._async_load_directory(item_path, item_id))

    def _async_load_directory(self, path, parent):
        """异步加载目录内容"""
        try:
            # 显示加载状态
            self.status_bar.config(text=f"正在加载: {path.name}")

            # 获取排序后的目录和文件列表
            dirs = sorted([d for d in path.iterdir()
                           if d.is_dir() and not d.name.startswith('.')],
                          key=lambda x: x.name.lower())

            files = sorted([f for f in path.iterdir()
                            if f.is_file() and not f.name.startswith('.')],
                           key=lambda x: x.name.lower())

            # 先处理目录再处理文件
            for item in dirs:
                node = self.tree.insert(parent, 'end',
                                        text=item.name,
                                        values=[str(item)],
                                        open=False)
                # 添加临时子节点
                self.tree.insert(node, 'end', text="加载中...", values=["loading"])

            # 处理文件
            for item in files:
                self.tree.insert(parent, 'end',
                                 text=item.name,
                                 values=[str(item)])

            # 恢复状态栏
            self.status_bar.config(text="就绪")
        except Exception as e:
            self.log_error(f"异步加载目录失败: {str(e)}")
            messagebox.showerror("目录错误", f"读取失败: {str(e)}")

    def _get_selected_files(self):
        """获取选中的文件列表"""
        selected_items = self.tree.selection()
        return [
            Path(self.tree.item(item, 'values')[0])
            for item in selected_items
            if Path(self.tree.item(item, 'values')[0]).is_file()
        ]

    def parse_response(self, api_response, code_lines):
        """解析含Markdown代码块的响应（最终修正版）"""
        try:
            # 1. 解析外层API响应
            response_data = json.loads(api_response)

            # 检查响应是否完整
            if response_data["choices"][0]["finish_reason"] == "length":
                raise ValueError("API响应被截断，请尝试减少代码量或增加max_tokens参数")

            content = response_data["choices"][0]["message"]["content"]

            # 2. 提取Markdown代码块
            json_str = content
            if "```json" in content:
                start = content.find("```json") + len("```json")
                end = content.rfind("```")
                json_str = content[start:end].strip()

            # 3. 解析漏洞数据
            vulnerabilities = json.loads(json_str)
            if not isinstance(vulnerabilities, list):
                vulnerabilities = [vulnerabilities]

            results = []
            for vuln in vulnerabilities:
                # 4. 处理键名前可能存在的空格
                vuln = {k.strip(): v for k, v in vuln.items()}

                # 5. 校验必填字段
                required_fields = ["文件路径", "行号", "风险等级", "漏洞类型", "详细描述"]
                for field in required_fields:
                    if field not in vuln:
                        raise ValueError(f"缺少必填字段: {field}")

                # 6. 处理行号字段类型
                raw_lines = vuln["行号"]
                if isinstance(raw_lines, int):
                    line_numbers = [raw_lines]
                elif isinstance(raw_lines, list):
                    line_numbers = raw_lines
                else:
                    line_numbers = []

                # 7. 构建漏洞数据
                vuln_data = {
                    "文件路径": vuln["文件路径"].strip(),
                    "行号": line_numbers,
                    "风险等级": vuln["风险等级"].strip(),
                    "漏洞类型": vuln["漏洞类型"].strip(),
                    "风险点": vuln.get("风险点", "").strip(),
                    "Payload": vuln.get("Payload", "").strip(),
                    "详细描述": vuln["详细描述"].strip(),
                    "修复建议": vuln.get("修复建议", "").strip()
                }

                # 验证行号是否有效
                valid_line_numbers = []
                for line in vuln_data["行号"]:
                    if 0 < line <= len(code_lines):
                        valid_line_numbers.append(line)
                    else:
                        self.log_error(f"无效行号: {line}（文件总行数: {len(code_lines)})")

                vuln_data["行号"] = valid_line_numbers

                # 过滤漏洞类型为"无"的结果
                if vuln_data["漏洞类型"].lower() != "无":
                    results.append(vuln_data)

            return results
        except Exception as e:
            error_msg = f"响应解析失败: {str(e)}\n原始响应内容:\n{api_response}"
            self.log_error(error_msg)
            return []

    def display_results(self, file_path, vulnerabilities):
        """显示结果入口方法"""
        self.root.after(0, self._safe_display_results, file_path, vulnerabilities)

    def analyze_file(self, file_path):
        """分析单个文件"""
        start_time = time.time()
        self.log_info(f"开始分析文件: {file_path.name}, 后缀: {file_path.suffix}")

        try:
            # 检查是否已取消分析
            if hasattr(self, 'auto_analysis_cancelled') and self.auto_analysis_cancelled:
                self.log_info(f"分析已取消，跳过文件: {file_path.name}")
                return

            # 检查文件是否存在
            if not file_path.exists():
                self.log_error(f"文件不存在: {file_path}")
                self.event_queue.put(('progress', 1, None))
                return

            # 打印调试信息
            print(f"[DEBUG] 开始分析文件: {file_path.name}, 后缀: {file_path.suffix}")

            # 特殊处理pom.xml文件
            if file_path.name.lower() == 'pom.xml' or file_path.suffix.lower() in ['.xml', '.pom']:
                print(f"[DEBUG] 检测到XML/POM文件: {file_path.name}，强制使用智能分块")
                self.log_info(f"检测到XML/POM文件: {file_path.name}，强制使用智能分块")

                # 读取文件内容
                try:
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        code = f.read()
                        code_lines = code.splitlines()

                    # 直接调用大文件分析方法，无需检查行数
                    self.analyze_large_file(file_path, code_lines, code)
                except Exception as e:
                    self.log_error(f"读取XML文件失败: {str(e)}", file_path)
                    self.event_queue.put(('progress', 1, None))
                return

            # 读取文件内容
            try:
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    code_lines = f.readlines()  # 读取所有行
                    code = ''.join(code_lines)  # 合并为完整代码
            except UnicodeDecodeError:
                # 尝试使用二进制模式读取，然后解码
                try:
                    with open(file_path, 'rb') as f:
                        binary_data = f.read()
                        code = binary_data.decode('utf-8', errors='replace')
                        code_lines = code.splitlines(True)
                    self.log_info(f"使用二进制模式成功读取文件: {file_path.name}")
                except Exception as e:
                    self.log_error(f"读取文件失败: {str(e)}", file_path)
                    self.event_queue.put(('progress', 1, None))
                    return
            except Exception as e:
                self.log_error(f"读取文件失败: {str(e)}", file_path)
                self.event_queue.put(('progress', 1, None))
                return

            # 再次检查是否已取消分析
            if hasattr(self, 'auto_analysis_cancelled') and self.auto_analysis_cancelled:
                self.log_info(f"分析已取消，跳过文件: {file_path.name}")
                return

            # 根据文件大小和类型决定分析方法
            if len(code_lines) > 1000 or file_path.suffix.lower() in ['.xml', '.pom', '.java', '.php']:
                self.analyze_large_file(file_path, code_lines, code)
            else:
                self.analyze_small_file(file_path, code_lines, code)
        except Exception as e:
            self.log_error(f"分析文件时出错: {str(e)}", file_path)
            self.event_queue.put(('progress', 1, None))

    def analyze_small_file(self, file_path, code_lines, code):
        """直接处理小文件，不进行分块"""
        try:
            # 获取文件扩展名
            file_ext = file_path.suffix.lower()

            # 检查是否已取消分析
            if hasattr(self, 'auto_analysis_cancelled') and self.auto_analysis_cancelled:
                print(f"[DEBUG] 分析已取消，停止处理 {file_path.name}")
                self.log_info(f"分析已取消，停止处理 {file_path.name}")
                self.event_queue.put(('progress', 1, None))
                return

            # 记录开始分析
            print(f"[DEBUG] 开始分析小文件: {file_path.name}")
            self.log_info(f"开始分析小文件: {file_path.name}")

            # 更新状态栏
            self.status_bar.config(text=f"正在分析: {file_path.name}")

            # 对于Java文件，即使是小文件也进行分块处理
            if file_ext == '.java':
                print(f"[DEBUG] 检测到Java文件: {file_path.name}，使用智能分块")
                self.log_info(f"检测到Java文件: {file_path.name}，使用智能分块")
                self.analyze_large_file(file_path, code_lines, code)
                return

            # 添加文件信息和上下文提示
            context_info = f"# 文件: {file_path.name}\n"
            context_info += f"# 行数: {len(code_lines)}\n\n"

            # 构建带上下文的代码
            code_with_context = context_info + code

            # 调用API分析文件
            response = self.call_deepseek_api(code_with_context, file_ext, file_path)

            # 检查是否已取消分析
            if hasattr(self, 'auto_analysis_cancelled') and self.auto_analysis_cancelled:
                return

            if response['status_code'] == 200:
                # 解析结果
                vulnerabilities = self.parse_response(response['text'], code_lines)

                # 确保文件路径正确
                for vuln in vulnerabilities:
                    vuln["文件路径"] = str(file_path)

                # 显示结果
                if vulnerabilities:
                    self.display_results(file_path, vulnerabilities)
                    print(f"[DEBUG] {file_path.name} 分析完成，发现 {len(vulnerabilities)} 个漏洞")
                    self.log_info(f"{file_path.name} 分析完成，发现 {len(vulnerabilities)} 个漏洞")
                else:
                    print(f"[DEBUG] {file_path.name} 分析完成，未发现漏洞")
                    self.log_info(f"{file_path.name} 分析完成，未发现漏洞")
            else:
                self.log_error(f"分析失败: {response['text'][:200]}", file_path)

            # 更新进度
            self.event_queue.put(('progress', 1, None))

        except Exception as e:
            self.log_error(f"小文件分析失败: {str(e)}", file_path)
            print(f"[ERROR] 小文件分析失败: {str(e)}")
            self.event_queue.put(('progress', 1, None))

    def analyze_large_file(self, file_path, code_lines, full_code):
        """智能分块处理大文件，根据代码结构进行分块"""
        try:
            # 获取文件扩展名
            file_ext = file_path.suffix.lower()

            # 特殊处理pom.xml文件
            if file_path.name.lower() == 'pom.xml' and not file_ext:
                file_ext = '.xml'  # 强制设置为.xml

            print(f"[DEBUG] 验证文件扩展名: {file_ext}, 文件名: {file_path.name}")

            # 检查是否已取消分析
            if hasattr(self, 'auto_analysis_cancelled') and self.auto_analysis_cancelled:
                print(f"[DEBUG] 分析已取消，停止处理 {file_path.name}")
                self.log_info(f"分析已取消，停止处理 {file_path.name}")
                self.event_queue.put(('progress', 1, None))
                return

            # 记录分块开始
            print(f"[DEBUG] 开始对 {file_path.name} 进行智能分块")
            self.log_info(f"开始对 {file_path.name} 进行智能分块")

            # 根据文件类型进行智能分块
            chunks = self._smart_code_chunking(full_code, file_ext)
            total_chunks = len(chunks)

            # 记录实际分块数量，但不再修改进度条最大值
            print(f"[DEBUG] {file_path.name} 被分为 {len(chunks)} 个代码块")
            self.log_info(f"{file_path.name} 被分为 {len(chunks)} 个代码块")

            # 更新状态栏
            self.status_bar.config(text=f"正在智能分块分析: {file_path.name} ({len(chunks)}个代码块)")

            # 初始化结果列表
            all_vulnerabilities = []

            # 支持多线程处理的文件类型
            multi_thread_exts = ['.xml', '.php', '.java', '.php']

            # 检查是否使用多线程处理
            use_multi_thread = False
            if file_ext.lower() in multi_thread_exts or file_path.name.lower() == 'pom.xml':
                use_multi_thread = True

            # 使用多线程处理支持的文件类型
            if use_multi_thread:
                print(f"[DEBUG] 使用多线程处理{file_ext}文件: {file_path.name}")
                self.log_info(f"使用多线程处理{file_ext}文件: {file_path.name}")

                # 使用多线程处理文件
                self._process_chunks_with_threads(file_path, chunks, file_ext, all_vulnerabilities)
                # 多线程处理完成后发送完成事件
                if not (hasattr(self, 'auto_analysis_cancelled') and self.auto_analysis_cancelled):
                    self.event_queue.put(('done', None, None))
            else:
                # 分块处理
                for i, (chunk, line_start, line_end, chunk_type) in enumerate(chunks):
                    # 检查是否已取消分析
                    if hasattr(self, 'auto_analysis_cancelled') and self.auto_analysis_cancelled:
                        print(f"[DEBUG] 分析已取消，停止处理 {file_path.name} 的剩余代码块")
                        self.log_info(f"分析已取消，停止处理 {file_path.name} 的剩余代码块")
                        break

                    # 检查是否暂停
                    if hasattr(self, 'auto_analysis_paused') and self.auto_analysis_paused:
                        while self.auto_analysis_paused and not self.auto_analysis_cancelled:
                            time.sleep(0.5)
                        # 再次检查是否已取消
                        if self.auto_analysis_cancelled:
                            break

                    # 记录当前处理的块
                    print(f"[DEBUG] 处理第 {i + 1}/{len(chunks)} 块: {chunk_type}, 行范围: {line_start}-{line_end}")
                    self.log_info(f"处理第 {i + 1}/{len(chunks)} 块: {chunk_type}, 行范围: {line_start}-{line_end}")

                    # 添加文件信息和上下文提示
                    context_info = f"# 文件: {file_path.name} (第{i + 1}/{len(chunks)}块)\n"
                    context_info += f"# 代码块类型: {chunk_type}\n"
                    context_info += f"# 行范围: {line_start}-{line_end}\n\n"

                    # 构建带上下文的代码块
                    chunk_with_context = context_info + chunk

                    # 调用API分析当前块
                    response = self.call_deepseek_api(chunk_with_context, file_ext, file_path)

                    # 再次检查是否已取消分析
                    if hasattr(self, 'auto_analysis_cancelled') and self.auto_analysis_cancelled:
                        break

                    if response['status_code'] == 200:
                        # 解析当前块的结果，并调整行号
                        chunk_vulnerabilities = self.parse_response(response['text'], chunk.splitlines())

                        # 调整行号，使其与原始文件对应
                        for vuln in chunk_vulnerabilities:
                            adjusted_lines = []
                            for line in vuln["行号"]:
                                adjusted_line = line_start + line - 1  # 减1是因为chunk的行号从1开始
                                adjusted_lines.append(adjusted_line)
                            vuln["行号"] = adjusted_lines
                            # 确保文件路径正确
                            vuln["文件路径"] = str(file_path)

                            # 将单个漏洞添加到总列表
                            all_vulnerabilities.append(vuln)

                            # 使用display_results方法来显示单个漏洞
                            self.display_results(file_path, [vuln])

                        # 记录块分析结果
                        print(f"[DEBUG] 第 {i + 1} 块分析完成，发现 {len(chunk_vulnerabilities)} 个漏洞")
                        self.log_info(f"第 {i + 1} 块分析完成，发现 {len(chunk_vulnerabilities)} 个漏洞")

                        # 更新进度
                        self.event_queue.put(('progress', 1, None))
                    else:
                        self.log_error(f"块{i + 1}分析失败: {response['text'][:200]}", file_path)
                        # 即使失败也更新进度
                        self.event_queue.put(('progress', 1, None))

            # 检查是否已取消分析，只有未取消时才发送结果
            if not (hasattr(self, 'auto_analysis_cancelled') and self.auto_analysis_cancelled) and all_vulnerabilities:
                # 不再需要发送完整结果，因为已经逐个发送了
                # self.event_queue.put(('result', (file_path, all_vulnerabilities), None))
                print(f"[DEBUG] {file_path.name} 分析完成，共发现 {len(all_vulnerabilities)} 个漏洞")
                self.log_info(f"{file_path.name} 分析完成，共发现 {len(all_vulnerabilities)} 个漏洞")

        except Exception as e:
            self.log_error(f"智能分块分析失败: {str(e)}", file_path)
            print(f"[ERROR] 智能分块分析失败: {str(e)}")
            self.event_queue.put(('progress', 1, None))
            self.event_queue.put(('done', None, None))  # 确保异常时发送完成事件

    def _handle_completed_future(self, future, futures, file_path):
        """处理已完成的任务"""
        # 查找对应的任务信息
        for f, (i, chunk_info) in futures:
            if f == future:
                chunk, line_start, line_end, chunk_type = chunk_info
                try:
                    # 获取结果
                    result = future.result()
                    if result:
                        if isinstance(result, dict) and 'status' in result:
                            if result['status'] == 'success':
                                print(f"[DEBUG] 线程处理完成第 {i + 1}/{len(futures)} 块")
                            else:
                                print(f"[DEBUG] 线程处理第 {i + 1}/{len(futures)} 块失败: {result['status']}")
                                print(f"[DEBUG] 错误信息: {result.get('error', '未知错误')}")
                        else:
                            print(f"[DEBUG] 线程处理完成第 {i + 1}/{len(futures)} 块")
                    else:
                        print(f"[DEBUG] 线程处理第 {i + 1}/{len(futures)} 块失败或无结果")
                except Exception as e:
                    print(f"[ERROR] 线程处理第 {i + 1}/{len(futures)} 块异常: {str(e)}")
                    self.log_error(f"线程处理第 {i + 1}/{len(futures)} 块异常: {str(e)}", file_path)
                break

    def _process_chunks_with_threads(self, file_path, chunks, file_ext, all_vulnerabilities):
        """使用多线程处理代码块"""
        import concurrent.futures
        from threading import Lock
        import os
        import time
        import traceback

        # 记录开始时间
        start_time = time.time()
        self.log_info(f"开始多线程处理文件: {file_path.name}, 代码块数量: {len(chunks)}")

        # 创建线程锁，用于保护共享资源
        result_lock = Lock()

        # 最大线程数 - 根据CPU核心数和块数量动态调整
        max_workers = min(5, os.cpu_count() or 4, len(chunks))

        print(f"[DEBUG] 启动多线程处理，最大线程数: {max_workers}")
        self.log_info(f"启动多线程处理，最大线程数: {max_workers}")

        # 创建进度跟踪变量
        processed_chunks = 0
        successful_chunks = 0
        failed_chunks = 0

        # 定义处理单个块的函数
        def process_chunk(chunk_info):
            i, (chunk, line_start, line_end, chunk_type) = chunk_info
            chunk_start_time = time.time()

            try:
                # 检查是否已取消分析
                if hasattr(self, 'auto_analysis_cancelled') and self.auto_analysis_cancelled:
                    self.log_info(f"分析已取消，跳过块 {i + 1}/{len(chunks)}")
                    return None

                # 检查是否暂停
                if hasattr(self, 'auto_analysis_paused') and self.auto_analysis_paused:
                    self.log_info(f"分析已暂停，块 {i + 1}/{len(chunks)} 等待继续")
                    while self.auto_analysis_paused and not self.auto_analysis_cancelled:
                        time.sleep(0.5)
                    # 再次检查是否已取消
                    if self.auto_analysis_cancelled:
                        self.log_info(f"分析已取消，跳过块 {i + 1}/{len(chunks)}")
                        return None
                    self.log_info(f"分析继续，处理块 {i + 1}/{len(chunks)}")

                # 记录当前处理的块
                print(f"[DEBUG] 线程处理第 {i + 1}/{len(chunks)} 块: {chunk_type}, 行范围: {line_start}-{line_end}")
                self.log_info(f"线程处理第 {i + 1}/{len(chunks)} 块: {chunk_type}, 行范围: {line_start}-{line_end}")

                # 添加文件信息和上下文提示
                context_info = f"# 文件: {file_path.name} (第{i + 1}/{len(chunks)}块)\n"
                context_info += f"# 代码块类型: {chunk_type}\n"
                context_info += f"# 行范围: {line_start}-{line_end}\n\n"

                # 构建带上下文的代码块
                chunk_with_context = context_info + chunk

                # 调用API分析当前块
                try:
                    response = self.call_deepseek_api(chunk_with_context, file_ext, file_path)
                except Exception as e:
                    error_msg = f"API调用失败，块 {i + 1}/{len(chunks)}: {str(e)}"
                    self.log_error(f"{error_msg}\n{traceback.format_exc()}", file_path)
                    print(f"[DEBUG] 线程处理第 {i + 1}/{len(chunks)} 块API调用失败: {str(e)}")
                    # 即使失败也更新进度
                    self.event_queue.put(('progress', 1, None))
                    return {'status': 'api_error', 'error': error_msg, 'chunk_index': i}

                # 再次检查是否已取消分析
                if hasattr(self, 'auto_analysis_cancelled') and self.auto_analysis_cancelled:
                    self.log_info(f"分析已取消，跳过结果处理: 块 {i + 1}/{len(chunks)}")
                    return None

                if response['status_code'] == 200:
                    try:
                        # 解析当前块的结果，并调整行号
                        chunk_vulnerabilities = self.parse_response(response['text'], chunk.splitlines())

                        # 调整行号，使其与原始文件对应
                        for vuln in chunk_vulnerabilities:
                            adjusted_lines = []
                            for line in vuln["行号"]:
                                adjusted_line = line_start + line - 1  # 减1是因为chunk的行号从1开始
                                adjusted_lines.append(adjusted_line)
                            vuln["行号"] = adjusted_lines
                            # 确保文件路径正确
                            vuln["文件路径"] = str(file_path)

                        # 记录块分析结果
                        print(
                            f"[DEBUG] 线程完成第 {i + 1}/{len(chunks)} 块分析，发现 {len(chunk_vulnerabilities)} 个漏洞")
                        self.log_info(
                            f"线程完成第 {i + 1}/{len(chunks)} 块分析，发现 {len(chunk_vulnerabilities)} 个漏洞")

                        # 使用锁保护共享资源
                        with result_lock:
                            # 将漏洞添加到总列表
                            all_vulnerabilities.extend(chunk_vulnerabilities)

                            # 更新UI显示每个漏洞
                            for vuln in chunk_vulnerabilities:
                                try:
                                    self.display_results(file_path, [vuln])
                                except Exception as e:
                                    self.log_error(f"显示漏洞结果失败: {str(e)}", file_path)

                        # 更新进度
                        self.event_queue.put(('progress', 1, None))

                        # 记录处理时间
                        chunk_time = time.time() - chunk_start_time
                        print(f"[DEBUG] 线程处理完成第 {i + 1}/{len(chunks)} 块，耗时: {chunk_time:.2f}秒")

                        return {'status': 'success', 'vulnerabilities': chunk_vulnerabilities, 'chunk_index': i}
                    except Exception as e:
                        error_msg = f"解析响应失败，块 {i + 1}/{len(chunks)}: {str(e)}"
                        self.log_error(f"{error_msg}\n{traceback.format_exc()}", file_path)
                        print(f"[DEBUG] 线程处理第 {i + 1}/{len(chunks)} 块解析失败: {str(e)}")
                        # 即使失败也更新进度
                        self.event_queue.put(('progress', 1, None))
                        return {'status': 'parse_error', 'error': error_msg, 'chunk_index': i}
                else:
                    # 初始化错误消息变量，避免未定义错误
                    error_msg = f"线程块{i + 1}分析失败: HTTP错误 {response['status_code']}"

                    # 特殊处理401认证错误
                    if response['status_code'] == 401:
                        error_message = "API密钥无效，请检查API密钥是否正确"
                        # 尝试从响应中提取错误消息
                        try:
                            response_json = json.loads(response['text'])
                            if 'error' in response_json and 'message' in response_json['error']:
                                error_message = response_json['error']['message']
                        except:
                            pass

                        # 只在未显示过错误时显示
                        if not self.api_validation_error_shown:
                            error_msg = f"API认证失败: {error_message}"
                            self.api_validation_error_shown = True
                            self.log_error(error_msg, file_path)

                            # 在UI线程中显示错误并取消分析
                            self.root.after(0, lambda message=error_message: [
                                messagebox.showerror("API认证失败", f"API认证失败: {message}"),
                                self.status_bar.config(text=f"API认证失败: {message}"),
                                setattr(self, 'auto_analysis_cancelled', True),
                                setattr(self, 'api_validated', False)
                            ])
                    elif response['status_code'] == 500:
                        # 处理服务器内部错误，不显示弹窗，只记录日志和更新状态栏
                        try:
                            error_text = response['text'][:200]
                            error_msg = f"API服务器内部错误: {error_text}"
                        except:
                            error_msg = "API服务器内部错误"

                        self.log_error(error_msg, file_path)

                        # 只更新状态栏，不弹窗
                        self.root.after(0, lambda err=error_msg:
                        self.status_bar.config(text=f"API请求失败: 服务器内部错误"))
                    elif response['status_code'] == 408:
                        # 处理超时错误，不显示弹窗
                        error_msg = f"API请求超时: {response['text']}"
                        self.log_error(error_msg, file_path)

                        # 只更新状态栏，不弹窗
                        self.root.after(0, lambda:
                        self.status_bar.config(text="API请求超时，请稍后重试"))
                    elif response['status_code'] == 503:
                        # 处理连接错误，不显示弹窗
                        error_msg = f"API连接失败: {response['text']}"
                        self.log_error(error_msg, file_path)

                        # 只更新状态栏，不弹窗
                        self.root.after(0, lambda:
                        self.status_bar.config(text="API连接失败，请检查网络"))
                    else:
                        # 处理其他HTTP错误
                        try:
                            error_text = response['text'][:200]
                            error_msg = f"API请求失败: 状态码 {response['status_code']} - {error_text}"
                        except:
                            error_msg = f"API请求失败: 状态码 {response['status_code']}"

                        self.log_error(error_msg, file_path)

                        # 在UI线程中显示错误消息
                        self.root.after(0, lambda err=error_msg: [
                            self.status_bar.config(text=f"API请求失败: 状态码 {response['status_code']}"),
                            self.show_error(f"API请求失败: 状态码 {response['status_code']}")
                        ])

                    # 即使失败也更新进度
                    self.event_queue.put(('progress', 1, None))
                    return {'status': 'http_error', 'error': error_msg, 'status_code': response['status_code'],
                            'chunk_index': i}
            except Exception as e:
                error_msg = f"线程处理第 {i + 1}/{len(chunks)} 块异常: {str(e)}"
                self.log_error(f"{error_msg}\n{traceback.format_exc()}", file_path)
                print(f"[ERROR] 线程处理第 {i + 1}/{len(chunks)} 块异常: {str(e)}")
                # 确保异常情况下也更新进度
                self.event_queue.put(('progress', 1, None))
                return {'status': 'exception', 'error': error_msg, 'chunk_index': i}

        try:
            # 创建线程池
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                # 不再一次性提交所有任务，而是逐个提交并检查暂停状态
                futures = []
                future_to_chunk = {}  # 创建future到chunk的映射字典
                active_futures = set()  # 跟踪活动的future
                i = 0

                # 处理所有块，直到全部完成或取消
                while i < len(chunks) or active_futures:
                    # 检查是否已取消分析
                    if hasattr(self, 'auto_analysis_cancelled') and self.auto_analysis_cancelled:
                        self.log_info(f"分析已取消，停止提交任务")
                        break

                    # 检查是否暂停
                    if hasattr(self, 'auto_analysis_paused') and self.auto_analysis_paused:
                        self.log_info(f"分析已暂停，等待继续提交任务")
                        # 只等待已提交的任务完成，不提交新任务
                        if active_futures:
                            # 等待任意一个任务完成
                            done, active_futures = concurrent.futures.wait(
                                active_futures,
                                return_when=concurrent.futures.FIRST_COMPLETED
                            )

                            # 处理完成的任务
                            for future in done:
                                self._handle_completed_future(future, futures, file_path)
                        else:
                            # 如果没有活动任务，则等待继续
                            while self.auto_analysis_paused and not self.auto_analysis_cancelled:
                                time.sleep(0.5)
                            # 再次检查是否已取消
                            if self.auto_analysis_cancelled:
                                self.log_info(f"分析已取消，停止提交任务")
                                break
                            self.log_info(f"分析继续，继续提交任务")
                        continue

                    # 提交新任务（如果还有未处理的块）
                    if i < len(chunks) and len(active_futures) < max_workers:
                        chunk_info = chunks[i]
                        future = executor.submit(process_chunk, (i, chunk_info))
                        futures.append((future, (i, chunk_info)))
                        future_to_chunk[future] = (i, chunk_info)
                        active_futures.add(future)
                        i += 1
                        continue

                    # 如果没有新任务可提交或已达到最大工作线程数，等待任意一个任务完成
                    if active_futures:
                        done, active_futures = concurrent.futures.wait(
                            active_futures,
                            return_when=concurrent.futures.FIRST_COMPLETED
                        )

                        # 处理完成的任务
                        for future in done:
                            self._handle_completed_future(future, futures, file_path)
                            # 从future_to_chunk中移除已处理的future
                            if future in future_to_chunk:
                                del future_to_chunk[future]
                    else:
                        # 所有任务都已完成
                        break

                # 处理完成的任务
                completed = 0
                # 使用futures列表中的future进行处理，而不是future_to_chunk
                for future, (i, chunk_info) in futures:
                    chunk, line_start, line_end, chunk_type = chunk_info
                    completed += 1
                    processed_chunks += 1

                    try:
                        # 获取结果
                        result = future.result()
                        if result:
                            if result['status'] == 'success':
                                successful_chunks += 1
                                print(f"[DEBUG] 线程处理完成第 {i + 1}/{len(chunks)} 块")
                            else:
                                failed_chunks += 1
                                print(f"[DEBUG] 线程处理第 {i + 1}/{len(chunks)} 块失败: {result['status']}")
                                print(f"[DEBUG] 错误信息: {result.get('error', '未知错误')}")
                        else:
                            failed_chunks += 1
                            print(f"[DEBUG] 线程处理第 {i + 1}/{len(chunks)} 块失败或无结果")
                    except Exception as e:
                        failed_chunks += 1
                        print(f"[ERROR] 线程处理第 {i + 1}/{len(chunks)} 块异常: {str(e)}")
                        self.log_error(f"线程处理第 {i + 1}/{len(chunks)} 块异常: {str(e)}", file_path)

                    # 更新进度信息
                    progress_percent = int(completed / len(chunks) * 100)
                    self.log_info(f"多线程处理进度: {completed}/{len(chunks)} ({progress_percent}%)")
                    print(
                        f"[DEBUG] 进度：{completed}/{len(chunks)} ({progress_percent}%) [成功: {successful_chunks}, 失败: {failed_chunks}]")
        except Exception as e:
            self.log_error(f"线程池执行异常: {str(e)}\n{traceback.format_exc()}", file_path)
        finally:
            # 记录总耗时和成功/失败统计
            total_time = time.time() - start_time
            self.log_info(
                f"多线程处理完成: {file_path.name}, 总耗时: {total_time:.2f}秒, 总块数: {len(chunks)}, 成功: {successful_chunks}, 失败: {failed_chunks}, 发现漏洞: {len(all_vulnerabilities)}个")
            print(
                f"[DEBUG] 异常分块只有{successful_chunks}块，进度：{successful_chunks}/{len(chunks)} （{int(successful_chunks / len(chunks) * 100)}%）")

            # 确保更新进度
            self.event_queue.put(('progress', 1, None))

    def process_event_queue(self):
        """处理事件队列中的事件"""
        try:
            while True:
                event, data, callback = self.event_queue.get_nowait()

                if event == 'result':
                    file_path, vulnerabilities = data
                    self.update_vulnerability_list(file_path, vulnerabilities)

                elif event == 'partial_result':
                    # 处理增量更新
                    file_path, chunk_vulnerabilities = data
                    # 将新发现的漏洞添加到现有列表中
                    if file_path not in self.vulnerabilities:
                        self.vulnerabilities[file_path] = []
                    self.vulnerabilities[file_path].extend(chunk_vulnerabilities)
                    # 更新界面显示
                    self.update_vulnerability_treeview()

                elif event == 'single_vuln':
                    # 处理单个漏洞更新
                    file_path, vuln_list = data
                    # 确保文件路径存在于漏洞字典中
                    if file_path not in self.vulnerabilities:
                        self.vulnerabilities[file_path] = []
                    # 添加单个漏洞
                    self.vulnerabilities[file_path].extend(vuln_list)
                    # 更新界面显示
                    self.update_vulnerability_treeview()

                elif event == 'progress':
                    self.update_progress()

                elif event == 'error':
                    error_msg = data
                    self.show_error(error_msg)

                # 添加对'done'事件的处理，用于重置分析按钮状态
                elif event == 'done':
                    # 重置分析按钮状态
                    if hasattr(self, 'btn_analyze'):
                        self.btn_analyze.config(text="开始分析", command=self.start_analysis)
                    # 重置分析状态标志
                    if hasattr(self, 'auto_analysis_cancelled'):
                        self.auto_analysis_cancelled = False
                    if hasattr(self, 'auto_analysis_paused'):
                        self.auto_analysis_paused = False
                    # 更新状态栏
                    self.status_bar.config(text="分析完成")

                if callback:
                    callback()

                self.event_queue.task_done()
        except queue.Empty:  # 修改这里，使用queue.Empty而不是Empty
            pass

        # 继续轮询事件队列
        self.root.after(100, self.process_event_queue)

    def update_vulnerability_list(self, file_path, vulnerabilities):
        """更新漏洞列表"""
        # 存储漏洞信息
        self.vulnerabilities[file_path] = vulnerabilities

        # 更新界面显示
        self.update_vulnerability_treeview()

    def update_progress(self):
        """更新进度显示"""
        try:
            # 更新进度条
            if hasattr(self, 'progress') and self.progress:
                current_value = self.progress['value']
                self.progress['value'] = current_value + 1

                # 更新进度文本
                total = len(self.files_to_analyze) if hasattr(self, 'files_to_analyze') and self.files_to_analyze else 0
                current = int(self.progress['value'])
                max_value = int(self.progress['maximum'])

                if max_value > 0:
                    percentage = int(current / max_value * 100)
                    self.status_bar.config(text=f"进度: {current}/{max_value} ({percentage}%)")
        except Exception as e:
            self.log_error(f"更新进度失败: {str(e)}")

    def show_error(self, error_msg):
        """显示错误信息"""
        self.log_error(error_msg)
        messagebox.showerror("错误", error_msg)

    def update_vulnerability_treeview(self):
        """更新漏洞列表视图"""
        # 清空当前视图
        for item in self.result_tree.get_children():
            self.result_tree.delete(item)
        vuln_id = 1
        # 添加所有漏洞到视图
        for file_path, vulns in self.vulnerabilities.items():
            for vuln in vulns:
                # 获取行号字符串
                line_numbers = ", ".join(map(str, vuln["行号"]))

                # 获取风险等级并标准化
                risk_level = vuln["风险等级"].strip()
                tag = None

                # 根据风险等级设置标签
                if "高" in risk_level:
                    tag = "高危"
                elif "中" in risk_level:
                    tag = "中危"
                elif "低" in risk_level:
                    tag = "低危"
                elif "提示" in risk_level or "信息" in risk_level:
                    tag = "提示"

                # 插入到树形视图
                item_id = self.result_tree.insert("", "end", values=(
                    str(vuln_id),
                    vuln["漏洞类型"],
                    vuln["风险等级"],
                    str(file_path),
                    vuln["详细描述"],
                    vuln["风险点"],
                    vuln["Payload"],
                    vuln["修复建议"],
                    line_numbers
                ))

                # 应用标签（如果有）
                if tag:
                    self.result_tree.item(item_id, tags=(tag,))

                vuln_id += 1

        # 更新状态栏
        total_vulns = sum(len(vulns) for vulns in self.vulnerabilities.values())
        self.status_bar.config(text=f"共发现 {total_vulns} 个漏洞")

    def log_info(self, message, file_path=None):
        """记录信息日志"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"[INFO][{timestamp}] {message}"

        if file_path:
            log_entry += f" - {file_path.name}"

        # 将日志写入文件
        with open("deepaudit_log.txt", "a", encoding="utf-8") as log_file:
            log_file.write(log_entry + "\n")

        # 如果有日志窗口，也可以显示在界面上
        if hasattr(self, 'log_text'):
            self.log_text.config(state='normal')
            self.log_text.insert(tk.END, log_entry + "\n")
            self.log_text.see(tk.END)
            self.log_text.config(state='disabled')

    def create_log_panel(self):
        """创建日志面板"""
        log_frame = ttk.LabelFrame(self.root, text="日志")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 创建日志文本框和滚动条
        self.log_text = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, height=8)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log_text.config(state='disabled')

        # 添加清除按钮
        clear_btn = ttk.Button(log_frame, text="清除日志", command=self.clear_log)
        clear_btn.pack(side=tk.RIGHT, padx=5, pady=5)

    def clear_log(self):
        """清除日志"""
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state='disabled')

    def _smart_code_chunking(self, code, file_ext):
        """根据代码结构智能分块"""
        chunks = []
        lines = code.splitlines()

        # 打印调试信息
        print(f"[DEBUG] 智能分块处理文件类型: {file_ext}")

        # 根据文件类型选择不同的分块策略
        if file_ext in ['.java']:
            print(f"[DEBUG] 调用Java分块处理: {file_ext}")
            chunks = self._chunk_java_code(lines)
            print(f"[DEBUG] Java分块完成，共 {len(chunks)} 个代码块")
        elif file_ext in ['.php']:
            print(f"[DEBUG] 调用PHP分块处理: {file_ext}")
            chunks = self._chunk_php_code(lines)
            print(f"[DEBUG] PHP分块完成，共 {len(chunks)} 个代码块")
        elif file_ext.lower() in ['.xml', '.pom'] or 'pom.xml' in file_ext.lower():
            print(f"[DEBUG] 调用XML分块处理: {file_ext}")
            chunks = self._chunk_xml_code(lines)
            print(f"[DEBUG] XML分块完成，共 {len(chunks)} 个代码块")
        else:
            # 默认按固定大小分块
            chunk_size = 200  # 每块约200行
            for i in range(0, len(lines), chunk_size):
                end = min(i + chunk_size, len(lines))
                chunk_code = '\n'.join(lines[i:end])
                chunks.append((chunk_code, i + 1, end, "固定大小块"))

        # 确保所有分块按照起始行号排序
        chunks = sorted(chunks, key=lambda x: x[1])

        return chunks

    def _chunk_xml_code(self, lines):
        """XML/POM代码智能分块"""
        try:
            # 如果文件为空，直接返回
            if not lines:
                return [("", 1, 1, "空文件")]

            # 首先过滤掉XML注释行
            filtered_lines = []
            original_to_filtered = {}  # 原始行号到过滤后行号的映射
            filtered_to_original = {}  # 过滤后行号到原始行号的映射

            for i, line in enumerate(lines):
                # 跳过XML注释行
                if '<!--' in line and '-->' in line:
                    # 如果注释不是整行，保留非注释部分
                    parts = []
                    current_pos = 0
                    while current_pos < len(line):
                        comment_start = line.find('<!--', current_pos)
                        if comment_start == -1:
                            parts.append(line[current_pos:])
                            break

                        # 添加注释前的内容
                        if comment_start > current_pos:
                            parts.append(line[current_pos:comment_start])

                        # 跳过注释
                        comment_end = line.find('-->', comment_start) + 3
                        current_pos = comment_end

                    # 如果有非注释内容，添加到过滤后的行
                    non_comment_line = ''.join(parts).strip()
                    if non_comment_line:
                        filtered_lines.append(non_comment_line)
                        original_to_filtered[i] = len(filtered_lines) - 1
                        filtered_to_original[len(filtered_lines) - 1] = i
                elif '<!--' in line:
                    # 多行注释开始，跳过
                    in_comment = True
                    # 检查是否有注释前的内容
                    comment_start = line.find('<!--')
                    if comment_start > 0:
                        non_comment_part = line[:comment_start].strip()
                        if non_comment_part:
                            filtered_lines.append(non_comment_part)
                            original_to_filtered[i] = len(filtered_lines) - 1
                            filtered_to_original[len(filtered_lines) - 1] = i
                elif '-->' in line:
                    # 多行注释结束
                    in_comment = False
                    # 检查是否有注释后的内容
                    comment_end = line.find('-->') + 3
                    if comment_end < len(line):
                        non_comment_part = line[comment_end:].strip()
                        if non_comment_part:
                            filtered_lines.append(non_comment_part)
                            original_to_filtered[i] = len(filtered_lines) - 1
                            filtered_to_original[len(filtered_lines) - 1] = i
                else:
                    # 正常行，添加到过滤后的行
                    filtered_lines.append(line)
                    original_to_filtered[i] = len(filtered_lines) - 1
                    filtered_to_original[len(filtered_lines) - 1] = i

            # 首先识别XML声明
            chunks = []
            for i, line in enumerate(filtered_lines):
                if line.strip().startswith("<?xml"):
                    # 使用原始行号
                    original_i = filtered_to_original[i]
                    chunks.append((line.strip(), original_i + 1, original_i + 1, "XML声明"))
                    break

            # 使用正则表达式识别主要标签
            import re

            # 识别project标签的开始和结束
            project_start = None
            project_end = None
            for i, line in enumerate(filtered_lines):
                if re.search(r'<project\b', line) and project_start is None:
                    project_start = i
                if re.search(r'</project>', line):
                    project_end = i

            if project_start is not None and project_end is not None:
                # 识别主要子标签
                main_tags = ["modules", "properties", "dependencies", "dependencyManagement",
                             "build", "profiles", "parent", "distributionManagement"]

                # 记录已处理的行范围，避免重复
                processed_ranges = set()

                for tag in main_tags:
                    tag_start = None
                    tag_end = None
                    tag_depth = 0

                    for i in range(project_start, project_end + 1):
                        line = filtered_lines[i]

                        # 识别标签开始
                        if re.search(f'<{tag}\\b', line) and tag_start is None:
                            tag_start = i
                            tag_depth = 1

                        # 如果已找到开始标签，计算嵌套深度
                        if tag_start is not None and tag_end is None:
                            # 计算当前行中的开始和结束标签
                            starts = len(re.findall(f'<{tag}\\b', line))
                            ends = len(re.findall(f'</{tag}>', line))

                            # 更新深度
                            tag_depth += starts - ends

                            # 如果深度回到0，说明找到了结束标签
                            if tag_depth == 0:
                                tag_end = i

                                # 创建这个标签的块 - 使用原始行号
                                original_start = filtered_to_original[tag_start]
                                original_end = filtered_to_original[tag_end]

                                # 获取原始代码内容（包括注释）
                                tag_content = '\n'.join(lines[original_start:original_end + 1])
                                chunks.append((tag_content, original_start + 1, original_end + 1, f"{tag}配置"))

                                # 记录已处理的行范围
                                for j in range(tag_start, tag_end + 1):
                                    processed_ranges.add(j)

                # 处理单个dependency标签（只有在dependencies标签外的独立dependency才处理）
                i = project_start
                while i <= project_end:
                    if i in processed_ranges:
                        i += 1
                        continue

                    line = filtered_lines[i]
                    if re.search(r'<dependency\b', line) and not re.search(r'</dependencies>', line):
                        dep_start = i
                        dep_depth = 1

                        # 寻找dependency结束
                        j = i + 1
                        while j <= project_end:
                            if j >= len(filtered_lines):
                                break

                            dep_line = filtered_lines[j]

                            # 计算当前行中的开始和结束标签
                            starts = len(re.findall(r'<dependency\b', dep_line))
                            ends = len(re.findall(r'</dependency>', dep_line))

                            # 更新深度
                            dep_depth += starts - ends

                            # 如果深度回到0，说明找到了结束标签
                            if dep_depth == 0:
                                dep_end = j

                                # 创建这个依赖项的块 - 使用原始行号
                                original_start = filtered_to_original[dep_start]
                                original_end = filtered_to_original[dep_end]

                                # 获取原始代码内容（包括注释）
                                dep_content = '\n'.join(lines[original_start:original_end + 1])
                                chunks.append((dep_content, original_start + 1, original_end + 1, "依赖项"))

                                # 记录已处理的行范围
                                for k in range(dep_start, dep_end + 1):
                                    processed_ranges.add(k)

                                i = dep_end
                                break

                            j += 1

                    i += 1

                # 处理未被识别的部分（如modelVersion, groupId等）
                i = project_start
                while i <= project_end:
                    # 如果这一行已经被处理过，跳过
                    if i in processed_ranges:
                        i += 1
                        continue

                    # 找到一段连续的未处理行
                    start_unprocessed = i
                    while i <= project_end and i not in processed_ranges:
                        i += 1

                    # 如果找到了未处理的行段，创建一个块
                    if i > start_unprocessed:
                        # 转换为原始行号
                        original_start = filtered_to_original[start_unprocessed]
                        original_end = filtered_to_original[min(i - 1, len(filtered_to_original) - 1)]

                        # 获取原始代码内容（包括注释）
                        unprocessed_content = '\n'.join(lines[original_start:original_end + 1])
                        if unprocessed_content.strip():  # 确保内容不为空
                            chunks.append((unprocessed_content, original_start + 1, original_end + 1, "其他配置"))
            else:
                # 如果没有找到project标签，将整个文件作为一个块
                chunks.append(('\n'.join(lines), 1, len(lines), "完整XML文件"))

            # 过滤掉空内容的块
            chunks = [(code, start, end, chunk_type) for code, start, end, chunk_type in chunks if code.strip()]

            # 按行号排序
            chunks.sort(key=lambda x: x[1])

            return chunks

        except Exception as e:
            print(f"[ERROR] XML分块异常: {str(e)}")
            import traceback
            traceback.print_exc()
            return [('\n'.join(lines), 1, len(lines), "分块失败")]  # 返回完整代码作为单个块

    def _chunk_php_code(self, lines):
        """PHP代码智能分块"""
        chunks = []
        current_chunk = []
        current_type = "导入块"
        chunk_start_line = 1
        in_class = False
        in_function = False
        in_method = False
        brace_stack = []  # 使用栈来跟踪大括号匹配
        in_comment = False
        class_name = ""
        function_name = ""
        # 初始化这两个变量，避免未绑定错误
        method_brace_depth = 0
        function_brace_depth = 0

        for i, line in enumerate(lines):
            line_num = i + 1
            stripped = line.strip()

            # 处理多行注释
            if "/*" in line and "*/" not in line:
                in_comment = True
            if "*/" in line:
                in_comment = False

            # 跟踪大括号，忽略注释中的大括号
            if not in_comment and not stripped.startswith("//"):
                # 处理左大括号
                for char_pos, char in enumerate(line):
                    if char == '{':
                        # 检查是否在字符串内
                        if not self._is_in_string(line, char_pos):
                            brace_stack.append('{')
                    elif char == '}':
                        # 检查是否在字符串内
                        if not self._is_in_string(line, char_pos) and brace_stack:
                            brace_stack.pop()

            # 检测命名空间和use语句
            if re.match(r'^\s*(namespace|use)\s+', line) and not in_comment:
                if not current_chunk or current_type == "导入块":
                    current_chunk.append(line)
                    if current_type != "导入块":
                        current_type = "导入块"
                        chunk_start_line = line_num
                else:
                    chunks.append(('\n'.join(current_chunk), chunk_start_line, line_num - 1, current_type))
                    current_chunk = [line]
                    chunk_start_line = line_num
                    current_type = "导入块"
                continue

            # 检测类定义
            class_match = re.match(r'^\s*(abstract\s+|final\s+)?class\s+(\w+)', line)
            if not in_comment and class_match and not in_class:
                if current_chunk:
                    chunks.append(('\n'.join(current_chunk), chunk_start_line, line_num - 1, current_type))
                current_chunk = [line]
                chunk_start_line = line_num
                # 提取类名
                class_name = class_match.group(2) if class_match.group(2) else "匿名类"
                current_type = f"类定义({class_name})"
                in_class = True
                continue

            # 检测函数/方法定义
            function_match = re.match(r'^\s*(public\s+|private\s+|protected\s+|static\s+)*(function)\s+(\w+)', line)
            if not in_comment and function_match:
                function_name = function_match.group(3) if function_match.group(3) else "匿名函数"

                if in_class:
                    # 类方法
                    if current_chunk and not in_method and "类定义" not in current_type:
                        chunks.append(('\n'.join(current_chunk), chunk_start_line, line_num - 1, current_type))
                        current_chunk = []
                        chunk_start_line = line_num

                    # 添加方法定义行到当前块
                    current_chunk.append(line)
                    current_type = f"方法({class_name}.{function_name})"
                    in_method = True
                    # 记录方法开始时的大括号栈深度
                    method_brace_depth = len(brace_stack)
                    if '{' in line:
                        method_brace_depth += 1  # 如果当前行有左大括号，调整深度
                else:
                    # 独立函数
                    if current_chunk and not in_function:
                        chunks.append(('\n'.join(current_chunk), chunk_start_line, line_num - 1, current_type))
                        current_chunk = [line]
                        chunk_start_line = line_num
                    else:
                        current_chunk.append(line)

                    current_type = f"函数({function_name})"
                    in_function = True
                    # 记录函数开始时的大括号栈深度
                    function_brace_depth = len(brace_stack)
                    if '{' in line:
                        function_brace_depth += 1  # 如果当前行有左大括号，调整深度
                continue

            # 检测函数/方法结束 - 使用大括号栈来判断
            if in_method and '}' in line:
                current_chunk.append(line)
                # 检查当前行后大括号栈的深度是否回到方法开始前的水平
                if len(brace_stack) < method_brace_depth:
                    # 检查是否有连续的大括号结束（可能是嵌套方法或类的结束）
                    next_lines_have_method = False
                    next_method_line = 0
                    # 向前查看，寻找下一个方法定义或非空行
                    for j in range(i + 1, min(i + 20, len(lines))):
                        next_line = lines[j].strip()
                        # 跳过空行和注释行
                        if not next_line or next_line.startswith('//') or next_line.startswith('/*'):
                            continue

                        # 如果找到了方法定义
                        if re.match(r'^\s*(public\s+|private\s+|protected\s+|static\s+)*(function)\s+(\w+)', next_line):
                            next_lines_have_method = True
                            next_method_line = j
                            break

                        # 如果找到了非方法定义的代码行，将其包含在当前方法中
                        if not next_line.startswith('}'):
                            # 将这些行添加到当前块中
                            for k in range(i + 1, j + 1):
                                current_chunk.append(lines[k])
                            i = j  # 更新循环索引
                            break

                    # 如果找到了下一个方法定义，结束当前方法块
                    if next_lines_have_method:
                        # 方法结束
                        chunks.append(('\n'.join(current_chunk), chunk_start_line, next_method_line - 1, current_type))
                        current_chunk = []
                        chunk_start_line = next_method_line
                        current_type = f"类内代码({class_name})"
                        in_method = False
                        continue

            # 检测独立函数结束 - 类似的逻辑修改
            if in_function and not in_class and '}' in line:
                current_chunk.append(line)
                # 检查当前行后大括号栈的深度是否回到函数开始前的水平
                if len(brace_stack) < function_brace_depth:
                    # 检查是否有连续的大括号结束（可能是嵌套函数的结束）
                    next_lines_have_function = False
                    next_function_line = 0

                    # 向前查看，寻找下一个函数定义或非空行
                    for j in range(i + 1, min(i + 20, len(lines))):
                        next_line = lines[j].strip()
                        # 跳过空行和注释行
                        if not next_line or next_line.startswith('//') or next_line.startswith('/*'):
                            continue

                        # 如果找到了函数定义
                        if re.match(r'^\s*(function)\s+(\w+)', next_line):
                            next_lines_have_function = True
                            next_function_line = j
                            break

                        # 如果找到了非函数定义的代码行，将其包含在当前函数中
                        if not next_line.startswith('}'):
                            # 将这些行添加到当前块中
                            for k in range(i + 1, j + 1):
                                current_chunk.append(lines[k])
                            i = j  # 更新循环索引
                            break

                    # 如果找到了下一个函数定义，结束当前函数块
                    if next_lines_have_function:
                        # 函数结束
                        chunks.append(
                            ('\n'.join(current_chunk), chunk_start_line, next_function_line - 1, current_type))
                        current_chunk = []
                        chunk_start_line = next_function_line
                        current_type = "全局代码"
                        in_function = False
                        continue

            # 检测类结束
            if in_class and '}' in line and not in_method:
                current_chunk.append(line)
                # 检查是否是类的结束大括号
                if len(brace_stack) == 0:
                    # 检查下一个非空行是否是另一个类或函数的开始
                    next_class_or_function = False
                    next_line_index = 0

                    # 向前查看，寻找下一个类或函数定义或非空行
                    for j in range(i + 1, min(i + 20, len(lines))):
                        next_line = lines[j].strip()
                        # 跳过空行和注释行
                        if not next_line or next_line.startswith('//') or next_line.startswith('/*'):
                            continue

                        # 如果找到了类或函数定义
                        if re.match(r'^\s*(abstract\s+|final\s+)?class\s+(\w+)', next_line) or \
                                re.match(r'^\s*(function)\s+(\w+)', next_line):
                            next_class_or_function = True
                            next_line_index = j
                            break

                        # 如果找到了非类或函数定义的代码行，将其包含在当前类中
                        # 将这些行添加到当前块中
                        for k in range(i + 1, j + 1):
                            current_chunk.append(lines[k])
                        i = j  # 更新循环索引
                        break

                    # 如果找到了下一个类或函数定义，结束当前类块
                    if next_class_or_function:
                        chunks.append(('\n'.join(current_chunk), chunk_start_line, next_line_index - 1, current_type))
                        current_chunk = []
                        chunk_start_line = next_line_index
                        current_type = "全局代码"
                        in_class = False
                        class_name = ""
                        continue

            # 添加到当前块
            current_chunk.append(line)

        # 添加最后一个块
        if current_chunk:
            chunks.append(('\n'.join(current_chunk), chunk_start_line, len(lines), current_type))

        # 合并过小的块和注释块
        merged_chunks = []
        min_chunk_size = 50  # 最小块大小
        current_merged = None
        i = 0

        while i < len(chunks):
            chunk = chunks[i]
            code, start, end, chunk_type = chunk
            code_lines = code.splitlines()

            # 检查是否是注释块或者过小的块
            is_comment_block = all(
                line.strip().startswith("//") or line.strip().startswith("/*") or line.strip().startswith(
                    "*") or not line.strip() for line in code_lines)
            is_small_block = len(code_lines) < min_chunk_size

            # 如果是注释块或者过小的块，并且不是类定义或方法定义，考虑合并
            if (is_comment_block or is_small_block) and not chunk_type.startswith(
                    "类定义") and not chunk_type.startswith("方法") and not chunk_type.startswith("函数"):
                # 尝试与下一个块合并（如果是函数或方法）
                if i < len(chunks) - 1:
                    next_chunk = chunks[i + 1]
                    next_code, next_start, next_end, next_type = next_chunk

                    if next_type.startswith("函数") or next_type.startswith("方法"):
                        # 将当前注释块合并到下一个函数/方法块
                        merged_code = code + "\n" + next_code
                        merged_chunk = (merged_code, start, next_end, next_type)
                        chunks[i + 1] = merged_chunk
                        i += 1
                        continue

                # 尝试与上一个合并的块或当前块合并
                if current_merged is not None:
                    # 合并小块
                    merged_code, merged_start, merged_end, merged_type = current_merged
                    current_merged = (merged_code + "\n" + code, merged_start, end, f"{merged_type}+{chunk_type}")
                else:
                    current_merged = chunk
            else:
                if current_merged is not None:
                    merged_chunks.append(current_merged)
                    current_merged = None
                merged_chunks.append(chunk)

            i += 1

        # 在函数末尾，返回chunks前添加排序代码
        if current_merged is not None:
            merged_chunks.append(current_merged)

        # 按照起始行号排序分块
        sorted_chunks = sorted(merged_chunks if merged_chunks else chunks, key=lambda x: x[1])

        return sorted_chunks

    def _is_in_string(self, line, pos):
        """检查位置是否在字符串内"""
        # 简单实现：计算位置前的引号数量
        single_quotes = line[:pos].count("'") - line[:pos].count("\\'")
        double_quotes = line[:pos].count('"') - line[:pos].count('\\"')
        return (single_quotes % 2 == 1) or (double_quotes % 2 == 1)

    def _chunk_java_code(self, lines):
        """Java代码智能分块"""
        chunks = []
        current_chunk = []
        current_type = "全局代码"
        chunk_start_line = 1
        brace_stack = []  # 使用栈来跟踪大括号匹配
        in_class = False
        in_method = False
        in_comment = False
        in_javadoc = False  # 新增：标记是否在Javadoc注释中
        in_annotation = False
        class_name = ""
        method_name = ""
        pending_method = False
        pending_method_start = 0
        class_brace_depth = 0
        method_brace_depth = 0

        # 跳过所有导入语句和包声明
        non_import_start_line = 1
        import_lines = []
        for i, line in enumerate(lines):
            stripped = line.strip()
            # 收集包声明或导入语句
            if re.match(r'^\s*package\s+', line) or re.match(r'^\s*import\s+', line):
                import_lines.append(line)
                non_import_start_line = i + 2  # +2 是为了跳过当前行并从下一行开始
                continue
            # 找到第一个非导入语句，结束跳过
            if stripped and not stripped.startswith("//") and not stripped.startswith("/*"):
                break

        # 更新起始行号
        chunk_start_line = non_import_start_line

        # 处理剩余代码
        i = 0
        while i < len(lines):
            line = lines[i]
            line_num = i + 1

            # 跳过导入语句和包声明
            if re.match(r'^\s*package\s+', line) or re.match(r'^\s*import\s+', line):
                i += 1
                continue

            stripped = line.strip()

            # 处理Javadoc注释 - 新增逻辑
            if stripped.startswith("/**"):
                in_javadoc = True
                javadoc_start = i
                # 寻找Javadoc注释的结束
                while i < len(lines) and "*/" not in lines[i]:
                    i += 1
                if i < len(lines):  # 找到了结束标记
                    i += 1  # 跳过包含 */ 的行
                in_javadoc = False
                continue  # 跳过Javadoc注释，继续处理下一行

            # 处理普通多行注释
            if "/*" in line and "*/" not in line and not in_javadoc:
                in_comment = True
            if "*/" in line and not in_javadoc:
                in_comment = False

            # 处理注解
            if stripped.startswith("@") and not stripped.endswith(")"):
                in_annotation = True
            if in_annotation and ")" in line:
                in_annotation = False

            # 跟踪大括号，忽略注释中的大括号
            if not in_comment and not stripped.startswith("//"):
                # 处理左大括号
                for char_pos, char in enumerate(line):
                    if char == '{':
                        # 检查是否在字符串内
                        if not self._is_in_string(line, char_pos):
                            brace_stack.append('{')
                    elif char == '}':
                        # 检查是否在字符串内
                        if not self._is_in_string(line, char_pos) and brace_stack:
                            brace_stack.pop()

            # 检测类定义
            class_match = re.match(
                r'^\s*(public|private|protected)?\s*(static|final|abstract)?\s*(class|interface|enum)\s+(\w+)', line)
            if not in_comment and not in_annotation and class_match and not in_class:
                if current_chunk:
                    chunks.append(('\n'.join(current_chunk), chunk_start_line, line_num - 1, current_type))
                current_chunk = [line]
                chunk_start_line = line_num
                # 提取类名
                class_name = class_match.group(4) if class_match.group(4) else "匿名类"
                current_type = f"类定义({class_name})"
                in_class = True
                class_brace_depth = len(brace_stack)
                pending_method = False  # 重置待处理方法注释状态
                i += 1
                continue

            # 改进后的方法匹配模式，支持注解和多个修饰符
            method_match = re.match(
                r'^\s*((@\w+\s*(\([^)]*\))?\s+)+)?\s*((?:public|private|protected|static|final|abstract|synchronized|native|transient|volatile)\s+)*\s*(?:<[^>]+>\s+)?[\w.<>\[\],\s]+?\s+(\w+)\s*\(',
                line)

            # 如果没有匹配到标准方法定义，尝试匹配构造函数
            if not method_match and in_class:
                # 增强构造函数检测（支持注解和多个修饰符）
                constructor_match = re.match(
                    r'^\s*((@\w+\s+)*((public|private|protected|static|final|abstract|synchronized)\s+)*)*'
                    + re.escape(class_name) + r'\s*\(', line)
                if constructor_match:
                    method_match = constructor_match
                    method_name = class_name  # 构造函数名与类名相同

            if in_class and not in_comment and not in_annotation and method_match and not in_method:
                # 提取方法名
                if method_match.group(4) if len(method_match.groups()) >= 4 else None:
                    method_name = method_match.group(4)
                elif not method_name:  # 如果不是构造函数且没有提取到方法名
                    method_name = "匿名方法"

                # 如果当前块不是类定义的一部分，则创建新块
                if current_chunk and "类定义" not in current_type:
                    chunks.append(('\n'.join(current_chunk), chunk_start_line, line_num - 1, current_type))
                    current_chunk = []
                    chunk_start_line = line_num

                # 添加方法定义行到当前块
                current_chunk.append(line)
                current_type = f"方法({class_name}.{method_name})"
                in_method = True
                method_brace_depth = len(brace_stack)
                pending_method = False  # 重置待处理方法注释状态
                i += 1
                continue

            # 检测方法结束 - 改进的逻辑
            if in_method and '}' in line:
                current_chunk.append(line)

                # 检查大括号栈的深度是否回到方法开始前的水平
                # 严格匹配大括号层级（考虑嵌套代码块）
                if len(brace_stack) == method_brace_depth - 1 and re.search(r'^\s*}\s*$', line):
                    # 查找下一个方法定义
                    next_method_found = False
                    for j in range(i + 1, min(i + 20, len(lines))):
                        next_line = lines[j].strip()
                        # 跳过空行和注释行
                        if not next_line or next_line.startswith('//') or next_line.startswith('/*'):
                            continue

                        # 检查是否是方法定义
                        next_method_match = re.match(
                            r'^\s*(public|private|protected)?\s*(static|final|abstract|synchronized)?\s*(<.*>)?\s*[\w<>[\],\s\.]+\s+(\w+)\s*\(',
                            next_line)

                        if next_method_match or re.match(
                                r'^\s*(public|private|protected)?\s*' + re.escape(class_name) + r'\s*\(', next_line):
                            next_method_found = True
                            break

                        # 如果找到了非方法定义的实质性代码行，则不是方法结束
                        if next_line and not re.match(r'^\s*(}|@|//)', next_line):
                            break

                    # 如果确认是方法结束
                    if next_method_found or len(brace_stack) < method_brace_depth:
                        chunks.append(('\n'.join(current_chunk), chunk_start_line, line_num, current_type))
                        current_chunk = []
                        chunk_start_line = line_num + 1
                        current_type = f"类内代码({class_name})"
                        in_method = False
                        method_name = ""  # 重置方法名
                i += 1
                continue

            # 检测类结束
            if in_class and len(brace_stack) <= class_brace_depth and '}' in line:
                current_chunk.append(line)

                # 检查是否是单独的结束括号行
                if line.strip() == '}':
                    # 检查下一个非空行是否是另一个类或方法的开始
                    next_class_or_method = False
                    next_line_index = 0

                    # 向前查看，寻找下一个类或方法定义
                    for j in range(i + 1, min(i + 20, len(lines))):
                        next_line = lines[j].strip()
                        # 跳过空行和注释行
                        if not next_line or next_line.startswith('//') or next_line.startswith('/*'):
                            continue

                        # 如果找到了类或方法定义
                        if (re.match(
                                r'^(public|private|protected)?\s*(static|final|abstract)?\s*(class|interface|enum)\s+(\w+)',
                                next_line) or
                                re.match(
                                    r'^(public|private|protected)?\s*(static|final|abstract|synchronized)?\s*(<.*>)?\s*[\w<>[\],\s\.]+\s+(\w+)\s*\(',
                                    next_line)):
                            next_class_or_method = True
                            next_line_index = j
                            break

                        # 如果找到了非类或方法定义的代码行
                        break

                    # 如果找到了下一个类或方法定义，结束当前类块
                    if next_class_or_method:
                        chunks.append(('\n'.join(current_chunk), chunk_start_line, next_line_index - 1, current_type))
                        current_chunk = []
                        chunk_start_line = next_line_index
                        current_type = "全局代码"
                        in_class = False
                        class_name = ""
                    else:
                        # 如果没有找到下一个类或方法，将这个结束括号与前面的代码合并
                        chunks.append(('\n'.join(current_chunk), chunk_start_line, line_num, current_type))
                        current_chunk = []
                        chunk_start_line = line_num + 1
                        current_type = "全局代码"
                        in_class = False
                        class_name = ""
                else:
                    # 如果不是单独的结束括号行，按原逻辑处理
                    chunks.append(('\n'.join(current_chunk), chunk_start_line, line_num, current_type))
                    current_chunk = []
                    chunk_start_line = line_num + 1
                    current_type = "全局代码"
                    in_class = False
                    class_name = ""

                i += 1
                continue

            # 添加到当前块
            current_chunk.append(line)
            i += 1

        # 添加最后一个块
        if current_chunk:
            chunks.append(('\n'.join(current_chunk), chunk_start_line, len(lines), current_type))

        # 合并过小的块和注释块
        merged_chunks = []
        min_chunk_size = 20  # 减小最小块大小，避免合并太多方法
        current_merged = None
        i = 0

        while i < len(chunks):
            chunk = chunks[i]
            code, start, end, chunk_type = chunk
            code_lines = code.splitlines()

            # 检查是否是注释块或者过小的块
            is_comment_block = all(
                line.strip().startswith("//") or line.strip().startswith("/*") or line.strip().startswith(
                    "*") or not line.strip() for line in code_lines)
            is_small_block = len(code_lines) < min_chunk_size

            # 如果是注释块或者过小的块，并且不是类定义或方法定义，考虑合并
            if (is_comment_block or is_small_block) and not chunk_type.startswith(
                    "类定义") and not chunk_type.startswith("方法"):
                # 尝试与下一个块合并（如果是方法）
                if i < len(chunks) - 1:
                    next_chunk = chunks[i + 1]
                    next_code, next_start, next_end, next_type = next_chunk

                    if next_type.startswith("方法"):
                        # 将当前注释块合并到下一个方法块
                        merged_code = code + "\n" + next_code
                        merged_chunk = (merged_code, start, next_end, next_type)
                        chunks[i + 1] = merged_chunk
                        i += 1
                        continue

                # 尝试与上一个合并的块或当前块合并
                if current_merged is not None:
                    # 合并小块
                    merged_code, merged_start, merged_end, merged_type = current_merged
                    current_merged = (merged_code + "\n" + code, merged_start, end, f"{merged_type}+{chunk_type}")
                else:
                    current_merged = chunk
            else:
                if current_merged is not None:
                    merged_chunks.append(current_merged)
                    current_merged = None
                merged_chunks.append(chunk)

            i += 1

        if current_merged is not None:
            merged_chunks.append(current_merged)

        # 按照起始行号排序分块
        sorted_chunks = sorted(merged_chunks if merged_chunks else chunks, key=lambda x: x[1])

        # 在这里添加合并单独的结束括号块的代码
        final_chunks = []
        i = 0
        while i < len(sorted_chunks):
            chunk = sorted_chunks[i]
            code, start, end, chunk_type = chunk

            # 检查是否是单独的结束括号块
            if code.strip() == "}" and "类内代码" in chunk_type:
                # 尝试与前一个块合并
                if i > 0:
                    prev_chunk = sorted_chunks[i - 1]
                    prev_code, prev_start, prev_end, prev_type = prev_chunk

                    # 提取类名
                    class_name_match = re.search(r'\(([^)]+)\)', chunk_type)
                    if class_name_match:
                        class_name = class_name_match.group(1)

                        # 如果前一个块是同一个类的一部分，将结束括号合并到前一个块
                        if class_name in prev_type:
                            merged_code = prev_code + "\n" + code
                            merged_chunk = (merged_code, prev_start, end, prev_type)
                            if final_chunks:  # 确保final_chunks不为空
                                final_chunks[len(final_chunks) - 1] = merged_chunk
                            else:
                                final_chunks.append(merged_chunk)
                        else:
                            final_chunks.append(chunk)
                    else:
                        final_chunks.append(chunk)
                else:
                    final_chunks.append(chunk)
            else:
                final_chunks.append(chunk)

            i += 1

        # 返回最终合并后的块
        return final_chunks if final_chunks else sorted_chunks

    def check_api_balance(self, api_key=None):
        """查询API余额

        Args:
            api_key: 可选的API密钥，如不提供则使用当前配置的密钥

        Returns:
            dict: 包含余额信息的字典，如查询失败则包含错误信息
        """
        try:
            # 使用参数提供的API密钥或配置中的密钥
            key = api_key if api_key else self.api_key

            if not key:
                return {"status": "error", "message": "未配置API密钥"}

            # 设置请求
            url = "https://api.deepseek.com/user/balance"
            headers = {
                "Accept": "application/json",
                "Authorization": f"Bearer {key}"
            }

            # 发送请求
            response = requests.get(url, headers=headers, timeout=10)

            # 处理响应
            if response.status_code == 200:
                try:
                    data = response.json()
                    return {"status": "success", "data": data}
                except:
                    return {"status": "error", "message": "解析余额数据失败", "raw": response.text}
            elif response.status_code == 401:
                return {"status": "error", "message": "API密钥无效或已过期"}
            else:
                return {"status": "error", "message": f"服务器返回错误: {response.status_code}", "raw": response.text}

        except requests.RequestException as e:
            return {"status": "error", "message": f"网络请求失败: {str(e)}"}
        except Exception as e:
            return {"status": "error", "message": f"查询余额时出错: {str(e)}"}


if __name__ == '__main__':
    # 启用高DPI支持，使UI在高分辨率显示器上更清晰
    try:
        from ctypes import windll

        windll.shcore.SetProcessDpiAwareness(1)  # 设置DPI感知级别
    except:
        pass  # 在非Windows系统上可能会失败，忽略错误

    root = tk.Tk()
    # 配置字体缩放
    try:
        root.tk.call('tk', 'scaling', 2.0)  # 调整UI缩放比例为2.0(原来是1.5)

        # 设置默认字体大小
        default_font = tk.font.nametofont("TkDefaultFont")
        default_font.configure(size=12)

        text_font = tk.font.nametofont("TkTextFont")
        text_font.configure(size=12)

        fixed_font = tk.font.nametofont("TkFixedFont")
        fixed_font.configure(size=12)
    except:
        pass

    app = CodeAuditApp(root)
    root.mainloop()
